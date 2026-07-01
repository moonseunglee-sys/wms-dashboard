"""
가동률 비교: DB(picking_zone_daily) vs 종합실적_최종.xlsx

행 구조 (H-I 기준, 표준시간 행=10):
  +0  = 표준시간[hr]
  +11 = 실적시간[hr]
  가동률 = 실적시간 / 표준시간 × 100
"""
import os
import openpyxl
import psycopg2
from dotenv import load_dotenv
from datetime import date, timedelta

load_dotenv()

FINAL_PATH = r"C:\Users\이문승\Desktop\PJ_생산성지표\data\master\2026년 06월 피킹 종합실적_최종.xlsx"

ZONE_STD_ROW = {
    'H-I': 10, 'C-D': 34, 'A-P': 58, 'DPS': 106,
    'E-F': 190, 'J-K': 214, 'L': 238, 'B': 262, 'L/S': 286,
    'M-N': 355, 'S': 379, 'W': 519, 'R': 543,
}
ACT_OFFSET = 11  # 표준시간 행 + 11 = 실적시간 행

BASE_DATE = date(2026, 6, 1)
BASE_COL = 4  # 열 4 = 6/1


def date_to_col(d):
    return BASE_COL + (d - BASE_DATE).days


def load_excel_data():
    wb = openpyxl.load_workbook(FINAL_PATH, read_only=True, data_only=True)
    ws = wb.worksheets[2]

    # 날짜 헤더 파악 (행 5)
    date_cols = {}
    for c in range(BASE_COL, BASE_COL + 35):
        v = ws.cell(5, c).value
        if v is None:
            continue
        if isinstance(v, str) and '/' in v:
            try:
                m, d_ = v.split('/')
                dt = date(2026, int(m), int(d_))
                date_cols[dt] = c
            except Exception:
                pass
        elif hasattr(v, 'date'):
            date_cols[v.date()] = c
        elif isinstance(v, date):
            date_cols[v] = c

    # 6/1은 문자열 "6/1"이므로 별도 처리
    if BASE_DATE not in date_cols:
        date_cols[BASE_DATE] = BASE_COL

    excel_data = {}
    for zone, std_row in ZONE_STD_ROW.items():
        act_row = std_row + ACT_OFFSET
        for dt, col in date_cols.items():
            std_val = ws.cell(std_row, col).value
            act_val = ws.cell(act_row, col).value
            if std_val and act_val and float(std_val) > 0:
                eff = float(act_val) / float(std_val) * 100
                excel_data[(dt.strftime('%Y-%m-%d'), zone)] = {
                    'std': float(std_val),
                    'act': float(act_val),
                    'eff': eff,
                }

    wb.close()
    return excel_data


def load_db_data():
    url = os.getenv('SUPABASE_POOLER_URL') or os.getenv('SUPABASE_DB_URL')
    conn = psycopg2.connect(url)
    cur = conn.cursor()
    cur.execute("""
        SELECT work_date, zone,
               SUM(std_time_hr) AS std_hr,
               SUM(act_time_hr) AS act_hr
        FROM picking_zone_daily
        GROUP BY work_date, zone
        ORDER BY work_date, zone
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    db_data = {}
    for work_date, zone, std_hr, act_hr in rows:
        ds = work_date.strftime('%Y-%m-%d') if hasattr(work_date, 'strftime') else str(work_date)
        if std_hr and float(std_hr) > 0:
            eff = float(act_hr or 0) / float(std_hr) * 100
            db_data[(ds, zone)] = {
                'std': float(std_hr),
                'act': float(act_hr or 0),
                'eff': eff,
            }
    return db_data


def compare(excel_data, db_data):
    all_keys = sorted(set(excel_data) | set(db_data))
    gaps = []

    print(f"\n{'날짜':<12} {'구역':<6} {'Excel_eff':>10} {'DB_eff':>10} {'차이':>8} {'비고'}")
    print("-" * 60)

    for key in all_keys:
        ds, zone = key
        ex = excel_data.get(key)
        db = db_data.get(key)

        if ex and db:
            diff = db['eff'] - ex['eff']
            flag = ' ★' if abs(diff) > 1.0 else ''
            print(f"{ds:<12} {zone:<6} {ex['eff']:>10.1f}% {db['eff']:>10.1f}% {diff:>+8.2f}%{flag}")
            if abs(diff) > 1.0:
                gaps.append((ds, zone, ex['eff'], db['eff'], diff))
        elif ex:
            print(f"{ds:<12} {zone:<6} {ex['eff']:>10.1f}%  {'DB없음':>10}")
        elif db:
            print(f"{ds:<12} {zone:<6} {'Excel없음':>10}  {db['eff']:>10.1f}%")

    print(f"\n총 비교: {len(all_keys)}건, 차이 1% 초과: {len(gaps)}건")

    if gaps:
        print("\n=== 주요 불일치 (차이 > 1%) ===")
        for ds, zone, ex_eff, db_eff, diff in sorted(gaps, key=lambda x: abs(x[4]), reverse=True):
            print(f"  {ds} {zone:<6}: Excel={ex_eff:.1f}% DB={db_eff:.1f}% 차이={diff:+.2f}%")


if __name__ == '__main__':
    print("Excel 데이터 로드 중...")
    excel_data = load_excel_data()
    print(f"  → {len(excel_data)}건 로드 (구역×날짜)")

    print("DB 데이터 로드 중...")
    db_data = load_db_data()
    print(f"  → {len(db_data)}건 로드 (구역×날짜)")

    compare(excel_data, db_data)
