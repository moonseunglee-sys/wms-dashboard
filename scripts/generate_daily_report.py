# -*- coding: utf-8 -*-
"""
일일 데이터 업데이트 리포트 생성 (관리자 검토용)

wms_rpa.py 파이프라인(자동화+DB적재) 완료 후 실행 — 그날 DB에 반영된 내용을
요약하고, "특이사항"(관리자 후속 조치가 필요한 항목)을 모아 엑셀로 저장.

특이사항 체크는 함수 단위로 등록하는 플러그인 구조 — ANOMALY_CHECKS 리스트에
함수만 추가하면 자동으로 리포트에 포함됨. 현재 등록된 체크:
  - 공장도가 0원: 단가를 못 찾아 금액이 0으로 처리된 품목 (기준정보_공장도가.xlsx
    갱신 필요 신호). picking_automation_v2.py / inbound_automation.py 가
    data/temp/price_gaps_*.json 으로 이미 저장해둔 갭 정보를 읽어 집계.

새 특이사항 체크 추가하는 법:
  1. `def check_xxx(ctx: dict) -> list[dict]:` 함수 작성 (ctx["target"], ctx["cur"] 사용 가능)
  2. 반환값은 {"구분","브랜드","유형","품목수","영향수량","상세","권장조치"} 형태의 dict 리스트
  3. ANOMALY_CHECKS 리스트에 함수 추가

Usage:
  python scripts/generate_daily_report.py --date 2026-07-07
"""
import argparse
import json
import os
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
import psycopg2
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Windows 콘솔 기본 코드페이지(cp949)로는 일부 문자 출력 시 UnicodeEncodeError로
# 스크립트 자체가 죽음 (subprocess로 캡처되어 실행돼도 자식 프로세스 자체의
# stdout 인코딩 문제라 부모 쪽 encoding 지정과 무관하게 발생) — 2026-07-10 확인
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

BASE_DIR = Path(__file__).resolve().parent.parent
TEMP_DIR = BASE_DIR / "data" / "temp"
OUT_DIR  = BASE_DIR / "data" / "reports"
load_dotenv(BASE_DIR / ".env")

DB_URL = os.getenv("SUPABASE_POOLER_URL") or os.getenv("SUPABASE_DB_URL")
INBOUND_BRANDS = ["일룸", "데스커", "퍼시스", "3PL"]


# ─────────────────────────────────────────────────────────────────────
# 특이사항 체크 (플러그인) — 여기에 함수를 추가하면 리포트에 자동 반영됨
# ─────────────────────────────────────────────────────────────────────
def check_price_gaps(ctx: dict) -> list[dict]:
    """공장도가 0원 처리된 품목 — 기준정보_공장도가.xlsx 갱신 필요 신호"""
    target = ctx["target"]
    findings = []

    p = TEMP_DIR / f"price_gaps_picking_{target}.json"
    if p.exists():
        gaps = json.loads(p.read_text(encoding="utf-8"))
        for brand, items in gaps.items():
            findings.append(_price_gap_row("피킹", brand, items))

    for brand in INBOUND_BRANDS:
        p2 = TEMP_DIR / f"price_gaps_inbound_{brand}_{target}.json"
        if p2.exists():
            items = json.loads(p2.read_text(encoding="utf-8"))
            findings.append(_price_gap_row("입고", brand, items))

    return findings


def _price_gap_row(source: str, brand: str, items: dict) -> dict:
    n_items = len(items)
    total_qty = sum(v["qty"] for v in items.values())
    codes = list(items.keys())
    detail = ", ".join(codes[:5]) + (f" 외 {n_items-5}종" if n_items > 5 else "")
    return {
        "구분": source, "브랜드": brand, "유형": "공장도가 0원",
        "품목수": n_items, "영향수량": round(total_qty, 1),
        "상세": detail,
        "권장조치": "기준정보_공장도가.xlsx에 해당 품목 단가 등록 후 재실행 필요",
    }


ANOMALY_CHECKS = [check_price_gaps]


# ─────────────────────────────────────────────────────────────────────
# DB 요약 조회
# ─────────────────────────────────────────────────────────────────────
def summarize_picking(cur, target: date) -> dict:
    cur.execute("""
        SELECT COUNT(*), COALESCE(SUM(std_time_hr),0), COALESCE(SUM(act_time_hr),0),
               COALESCE(SUM(pick_amount),0), COALESCE(SUM(pick_box),0)
        FROM picking_zone_daily WHERE work_date = %s
    """, (target,))
    n_zones, std, act, amt, box = cur.fetchone()
    cur.execute("SELECT COUNT(*) FROM picking_worker_daily WHERE work_date = %s", (target,))
    n_workers = cur.fetchone()[0]
    return {
        "구역수": n_zones, "작업자수": n_workers,
        "표준시간": float(std), "실적시간": float(act),
        "피킹금액": float(amt), "피킹박스": int(box),
    }


def summarize_inbound(cur, target: date) -> list[dict]:
    cur.execute("""
        SELECT brand, qty_total, amt_total, pallets, hours
        FROM inbound_brand_daily WHERE work_date = %s ORDER BY brand
    """, (target,))
    return [
        {"브랜드": r[0], "입고수량": float(r[1]), "입고금액": float(r[2]),
         "파렛트": int(r[3]), "실적시간": float(r[4])}
        for r in cur.fetchall()
    ]


# ─────────────────────────────────────────────────────────────────────
# 엑셀 생성
# ─────────────────────────────────────────────────────────────────────
_THIN, _THICK = Side(style="thin"), Side(style="medium")
HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
WARN_FILL = PatternFill("solid", fgColor="FDECEA")
OK_FILL   = PatternFill("solid", fgColor="E8F5E9")
BAND_FILL = PatternFill("solid", fgColor="F2F7FC")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
BOLD_FONT = Font(bold=True, size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right", vertical="center")


def _hdr(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.fill, c.font, c.alignment = HDR_FILL, HDR_FONT, CENTER
    return c


def _border_block(ws, r1, r2, c1, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = Border(
                top=_THIN, left=_THIN,
                bottom=_THICK if r == r2 else _THIN,
                right=_THICK if c == c2 else _THIN,
            )


def build_summary_sheet(wb, target: date, picking: dict, inbound: list[dict]):
    ws = wb.create_sheet("요약", 0)
    ws.column_dimensions["A"].width = 16
    for col in "BCDEF":
        ws.column_dimensions[col].width = 14

    row = 1
    title = ws.cell(row=row, column=1, value=f"일일 데이터 업데이트 리포트 — {target}")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    title.font = Font(bold=True, size=13)
    ws.row_dimensions[row].height = 24
    row += 2

    # 피킹 요약
    ws.cell(row=row, column=1, value="피킹 실적").font = BOLD_FONT
    row += 1
    headers = ["구역수", "작업자수", "표준시간(h)", "실적시간(h)", "피킹금액", "피킹박스"]
    for j, h in enumerate(headers, 1):
        _hdr(ws, row, j, h)
    row += 1
    vals = [picking["구역수"], picking["작업자수"], round(picking["표준시간"], 1),
            round(picking["실적시간"], 1), round(picking["피킹금액"]), picking["피킹박스"]]
    for j, v in enumerate(vals, 1):
        ws.cell(row=row, column=j, value=v).alignment = RIGHT
    _border_block(ws, row - 1, row, 1, 6)
    row += 3

    # 입고 요약
    ws.cell(row=row, column=1, value="입고 실적 (브랜드별)").font = BOLD_FONT
    row += 1
    headers2 = ["브랜드", "입고수량", "입고금액", "파렛트", "실적시간(h)"]
    for j, h in enumerate(headers2, 1):
        _hdr(ws, row, j, h)
    hdr_row = row
    row += 1
    if inbound:
        for i, r in enumerate(inbound):
            fill = BAND_FILL if i % 2 == 0 else None
            vals2 = [r["브랜드"], round(r["입고수량"]), round(r["입고금액"]),
                     r["파렛트"], round(r["실적시간"], 1)]
            for j, v in enumerate(vals2, 1):
                c = ws.cell(row=row, column=j, value=v)
                c.alignment = CENTER if j == 1 else RIGHT
                if fill:
                    c.fill = fill
            row += 1
    else:
        ws.cell(row=row, column=1, value="데이터 없음")
        row += 1
    _border_block(ws, hdr_row, row - 1, 1, 5)


def build_anomaly_sheet(wb, target: date, findings: list[dict]):
    ws = wb.create_sheet("특이사항")
    cols = ["구분", "브랜드", "유형", "품목수", "영향수량", "상세", "권장조치"]
    widths = [8, 10, 16, 8, 10, 36, 40]
    for j, (h, w) in enumerate(zip(cols, widths), 1):
        _hdr(ws, 1, j, h)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 22

    if not findings:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cols))
        c = ws.cell(row=2, column=1, value=f"✓ {target} — 특이사항 없음")
        c.fill, c.font, c.alignment = OK_FILL, Font(bold=True, color="2E7D32"), CENTER
        ws.row_dimensions[2].height = 24
        _border_block(ws, 1, 2, 1, len(cols))
        return

    for i, f in enumerate(findings):
        row = i + 2
        vals = [f["구분"], f["브랜드"], f["유형"], f["품목수"], f["영향수량"], f["상세"], f["권장조치"]]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=j, value=v)
            c.fill = WARN_FILL
            c.alignment = LEFT if j in (6, 7) else (CENTER if j in (1, 2, 3) else RIGHT)
    _border_block(ws, 1, len(findings) + 1, 1, len(cols))


# ─────────────────────────────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--date", required=True, help="YYYY-MM-DD")
    ap.add_argument("--out",  help="출력 파일 경로 (기본: data/reports/일일리포트_MMDD.xlsx)")
    args = ap.parse_args()
    target = datetime.strptime(args.date, "%Y-%m-%d").date()

    if not DB_URL:
        raise SystemExit("SUPABASE_POOLER_URL/SUPABASE_DB_URL 미설정")

    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()
    picking = summarize_picking(cur, target)
    inbound = summarize_inbound(cur, target)

    ctx = {"target": target, "cur": cur}
    findings = []
    for check in ANOMALY_CHECKS:
        findings.extend(check(ctx))
    cur.close()
    conn.close()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    build_summary_sheet(wb, target, picking, inbound)
    build_anomaly_sheet(wb, target, findings)

    out_path = Path(args.out) if args.out else OUT_DIR / f"일일리포트_{target.strftime('%m%d')}.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))

    print(f"리포트 저장: {out_path}")
    print(f"  피킹: 구역 {picking['구역수']}개, 작업자 {picking['작업자수']}명")
    print(f"  입고: {len(inbound)}개 브랜드")
    if findings:
        print(f"  [특이사항] {len(findings)}건 발견:")
        for f in findings:
            print(f"    - [{f['구분']}/{f['브랜드']}] {f['유형']}: {f['품목수']}종, {f['상세']}")
    else:
        print("  특이사항 없음")


if __name__ == "__main__":
    main()
