# -*- coding: utf-8 -*-
"""
입고 생산성 멀티시트 리포트 생성 (아카이브 JSON 기반)

시트 구성
  [합계]     : 브랜드 × 날짜별 실적수량(가로) + 기간 합계(수량/금액/파렛트/시간)
  [MM.DD] × N: 각 날짜의 브랜드별 실적 + 입고유형 구성(정산 6유형)

데이터 출처: data/daily/<YYYY-MM>/inbound_<date>.json
  (scripts/inbound_batch_run.py 로 미리 아카이브 필요)

Usage:
  python scripts/generate_inbound_report.py --start 2026-07-01 --end 2026-07-06
  python scripts/generate_inbound_report.py --dates 2026-07-01 2026-07-03
  python scripts/generate_inbound_report.py --start 2026-07-01 --end 2026-07-06 --out 입고_7월1주.xlsx
"""
import argparse
import json
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR  = Path(__file__).parent.parent
DAILY_DIR = BASE_DIR / "data" / "daily"
OUT_DIR   = BASE_DIR / "data" / "reports"

BRAND_ORDER = ["일룸", "퍼시스", "데스커", "3PL"]
TYPE_ORDER  = [
    ("normal",  "정상입고"),
    ("return",  "반품입고"),
    ("certify", "정품화입고"),
    ("reentry", "재입고"),
    ("inspect", "검사이동,업체반송"),
    ("cut",     "CUT"),
]


# ── 데이터 로드 ─────────────────────────────────────────────────────────
def load_day(d: date) -> list[dict] | None:
    p = DAILY_DIR / d.strftime("%Y-%m") / f"inbound_{d}.json"
    if p.exists():
        return json.loads(p.read_text(encoding="utf-8"))
    return None


def load_all(dates: list[date]) -> dict[date, list[dict]]:
    result = {}
    for d in dates:
        rows = load_day(d)
        if rows:
            result[d] = rows
        else:
            print(f"  [SKIP] {d} — inbound_{d}.json 없음 (inbound_batch_run.py 먼저 실행)")
    return result


def by_brand(rows: list[dict]) -> dict[str, list[dict]]:
    out: dict[str, list[dict]] = {}
    for r in rows:
        out.setdefault(r["brand"], []).append(r)
    return out


# ── 스타일 헬퍼 ─────────────────────────────────────────────────────────
_THIN, _THICK = Side(style="thin"), Side(style="medium")

def _fill(hex_color: str):
    return PatternFill("solid", fgColor=hex_color)

HDR_FILL, SUB_FILL, SUM_FILL = _fill("1F4E79"), _fill("2F75B6"), _fill("D6E4F7")
BAND_FILL, WHITE_FILL        = _fill("F2F7FC"), _fill("FFFFFF")
HDR_FONT, BOLD_FONT, BASE_FONT = Font(bold=True, color="FFFFFF", size=10), Font(bold=True, size=10), Font(size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

def _hdr(ws, row, col, value, fill=None, font=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill, c.font, c.alignment = fill or HDR_FILL, font or HDR_FONT, CENTER
    return c

def _cell(ws, row, col, value=None, fmt=None, fill=None, font=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    c.font = font or BASE_FONT
    c.alignment = align or RIGHT
    return c

def _border_block(ws, first_row, last_row, first_col, last_col):
    for row in range(first_row, last_row + 1):
        for col in range(first_col, last_col + 1):
            ws.cell(row=row, column=col).border = Border(
                top=_THIN, left=_THIN,
                bottom=_THICK if row == last_row else _THIN,
                right=_THICK if col == last_col else _THIN,
            )


# ── 합계 시트 ───────────────────────────────────────────────────────────
def build_summary(wb: openpyxl.Workbook, data: dict[date, list[dict]]):
    ws = wb.create_sheet("합계", 0)
    dates = sorted(data.keys())

    COL_BRAND = 1
    COL_DATES = {d: 2 + i for i, d in enumerate(dates)}
    COL_QTY, COL_AMT, COL_PLT, COL_HR = (2 + len(dates) + i for i in range(4))
    TOTAL_COLS = COL_HR

    ws.row_dimensions[1].height = 24
    ws.column_dimensions["A"].width = 10
    _hdr(ws, 1, COL_BRAND, "브랜드")
    for d, col in COL_DATES.items():
        _hdr(ws, 1, col, f"{d.month}.{d.day}")
        ws.column_dimensions[get_column_letter(col)].width = 10
    for col, label in [(COL_QTY, "기간\n수량"), (COL_AMT, "기간\n금액(백만)"),
                       (COL_PLT, "기간\n파렛트"), (COL_HR, "기간\n실적(h)")]:
        _hdr(ws, 1, col, label, fill=SUB_FILL)
        ws.column_dimensions[get_column_letter(col)].width = 12

    brand_by_date = {d: by_brand(rows) for d, rows in data.items()}
    active_brands = [b for b in BRAND_ORDER
                     if any(b in brand_by_date[d] for d in dates)]

    for i, brand in enumerate(active_brands):
        row = i + 2
        fill = BAND_FILL if i % 2 == 0 else WHITE_FILL
        _cell(ws, row, COL_BRAND, brand, fill=fill, font=BOLD_FONT, align=CENTER)

        tot_qty = tot_amt = tot_plt = tot_hr = 0.0
        for d, col in COL_DATES.items():
            brows = brand_by_date[d].get(brand, [])
            qty = sum(r["qty_total"] for r in brows)
            _cell(ws, row, col, round(qty) if qty else None, fmt='#,##0', fill=fill)
            tot_qty += qty
            tot_amt += sum(r["amt_total"] for r in brows)
            tot_plt += sum(r["basic_pallets"] for r in brows)
            tot_hr  += sum(r["hours"] for r in brows)

        _cell(ws, row, COL_QTY, round(tot_qty) if tot_qty else None, fmt='#,##0', fill=fill)
        _cell(ws, row, COL_AMT, round(tot_amt / 1_000_000, 1) if tot_amt else None, fmt='#,##0.0', fill=fill)
        _cell(ws, row, COL_PLT, round(tot_plt) if tot_plt else None, fmt='#,##0', fill=fill)
        _cell(ws, row, COL_HR,  round(tot_hr, 1) if tot_hr else None, fmt='0.0', fill=fill)

    sum_row = 2 + len(active_brands)
    _cell(ws, sum_row, COL_BRAND, "합계", fill=SUM_FILL, font=BOLD_FONT, align=CENTER)
    for d, col in COL_DATES.items():
        total = sum(r["qty_total"] for b in active_brands for r in brand_by_date[d].get(b, []))
        _cell(ws, sum_row, col, round(total) if total else None, fmt='#,##0', fill=SUM_FILL, font=BOLD_FONT)
    grand_qty = sum(r["qty_total"]      for d in dates for r in data[d])
    grand_amt = sum(r["amt_total"]      for d in dates for r in data[d])
    grand_plt = sum(r["basic_pallets"]  for d in dates for r in data[d])
    grand_hr  = sum(r["hours"]          for d in dates for r in data[d])
    _cell(ws, sum_row, COL_QTY, round(grand_qty), fmt='#,##0', fill=SUM_FILL, font=BOLD_FONT)
    _cell(ws, sum_row, COL_AMT, round(grand_amt / 1_000_000, 1), fmt='#,##0.0', fill=SUM_FILL, font=BOLD_FONT)
    _cell(ws, sum_row, COL_PLT, round(grand_plt), fmt='#,##0', fill=SUM_FILL, font=BOLD_FONT)
    _cell(ws, sum_row, COL_HR,  round(grand_hr, 1), fmt='0.0', fill=SUM_FILL, font=BOLD_FONT)

    _border_block(ws, 1, sum_row, 1, TOTAL_COLS)
    ws.freeze_panes = "B2"


# ── 개별 날짜 시트 ──────────────────────────────────────────────────────
def build_daily(wb: openpyxl.Workbook, d: date, rows: list[dict]):
    sheet_name = f"{d.month}.{d.day}"
    ws = wb.create_sheet(sheet_name)

    COLS = [
        ("브랜드", 10, None), ("입고수량", 12, '#,##0'), ("입고금액", 16, '#,##0'),
        ("파렛트", 10, '#,##0'), ("실적시간(h)", 12, '0.00'),
        ("시간당수량", 12, '#,##0'), ("시간당금액", 14, '#,##0'),
    ] + [(label, 13, '#,##0') for _, label in TYPE_ORDER]
    N = len(COLS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N)
    title = ws.cell(row=1, column=1, value=f"입고 생산성  {d.year}년 {d.month}월 {d.day}일")
    title.fill, title.font, title.alignment = _fill("1F4E79"), Font(bold=True, color="FFFFFF", size=12), CENTER
    ws.row_dimensions[1].height = 26

    ws.row_dimensions[2].height = 30
    for j, (name, width, _) in enumerate(COLS, 1):
        _hdr(ws, 2, j, name, fill=SUB_FILL)
        ws.column_dimensions[get_column_letter(j)].width = width

    grouped = by_brand(rows)
    active = [b for b in BRAND_ORDER if b in grouped]

    for i, brand in enumerate(active):
        row = i + 3
        brows = grouped[brand]
        qty = sum(r["qty_total"] for r in brows)
        amt = sum(r["amt_total"] for r in brows)
        plt = sum(r["basic_pallets"] for r in brows)
        hrs = sum(r["hours"] for r in brows)
        fill = BAND_FILL if i % 2 == 0 else WHITE_FILL

        values = [
            brand, round(qty), round(amt), round(plt), round(hrs, 4),
            round(qty / hrs) if hrs else None, round(amt / hrs) if hrs else None,
        ]
        for key, _ in TYPE_ORDER:
            values.append(round(sum(r.get(f"detail_qty_{key}", 0) for r in brows)) or None)

        for j, (val, (_, _, fmt)) in enumerate(zip(values, COLS), 1):
            align = CENTER if j == 1 else RIGHT
            _cell(ws, row, j, val, fmt=fmt, fill=fill,
                  font=BOLD_FONT if j == 1 else BASE_FONT, align=align)

    sum_row = 3 + len(active)
    tot_qty = sum(r["qty_total"] for b in active for r in grouped[b])
    tot_amt = sum(r["amt_total"] for b in active for r in grouped[b])
    tot_plt = sum(r["basic_pallets"] for b in active for r in grouped[b])
    tot_hr  = sum(r["hours"] for b in active for r in grouped[b])
    sum_vals = [
        "합계", round(tot_qty), round(tot_amt), round(tot_plt), round(tot_hr, 4),
        round(tot_qty / tot_hr) if tot_hr else None, round(tot_amt / tot_hr) if tot_hr else None,
    ]
    for key, _ in TYPE_ORDER:
        sum_vals.append(round(sum(r.get(f"detail_qty_{key}", 0) for b in active for r in grouped[b])) or None)
    for j, (val, (_, _, fmt)) in enumerate(zip(sum_vals, COLS), 1):
        align = CENTER if j == 1 else RIGHT
        _cell(ws, sum_row, j, val, fmt=fmt, fill=SUM_FILL, font=BOLD_FONT, align=align)

    _border_block(ws, 2, sum_row, 1, N)


# ── 메인 ──────────────────────────────────────────────────────────────
def date_range(start: date, end: date):
    d = start
    while d <= end:
        yield d
        d += timedelta(days=1)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--start", help="시작 날짜 YYYY-MM-DD")
    ap.add_argument("--end",   help="종료 날짜 YYYY-MM-DD")
    ap.add_argument("--dates", nargs="+", help="특정 날짜 목록 (--start/--end 대신)")
    ap.add_argument("--out",   help="출력 파일명 (기본: 입고생산성_MMDD_MMDD.xlsx)")
    args = ap.parse_args()

    if args.dates:
        dates = sorted(datetime.strptime(d, "%Y-%m-%d").date() for d in args.dates)
    elif args.start and args.end:
        dates = list(date_range(
            datetime.strptime(args.start, "%Y-%m-%d").date(),
            datetime.strptime(args.end, "%Y-%m-%d").date(),
        ))
    else:
        ap.error("--start/--end 또는 --dates 필요")

    print(f"\n리포트 생성 ({len(dates)}개 날짜)...")
    data = load_all(dates)
    if not data:
        print("데이터 없음. inbound_batch_run.py 로 먼저 아카이브 필요.")
        return

    loaded_dates = sorted(data.keys())
    print(f"  로드 완료: {loaded_dates[0]} ~ {loaded_dates[-1]} ({len(data)}일)")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    build_summary(wb, data)
    for d in loaded_dates:
        build_daily(wb, d, data[d])

    if args.out:
        out_path = Path(args.out)
        if not out_path.is_absolute():
            out_path = BASE_DIR / args.out
    else:
        s, e = loaded_dates[0].strftime("%m%d"), loaded_dates[-1].strftime("%m%d")
        fname = f"입고생산성_{s}_{e}.xlsx" if s != e else f"입고생산성_{s}.xlsx"
        out_path = OUT_DIR / fname

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f"\n저장 완료: {out_path}")
    print(f"  시트: 합계 + {len(loaded_dates)}개 일자")


if __name__ == "__main__":
    main()
