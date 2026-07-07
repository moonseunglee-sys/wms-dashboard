# -*- coding: utf-8 -*-
"""
피킹 생산성 멀티시트 리포트 생성

시트 구성
  [합계]     : 구역 × 날짜별 실적시간(가로) + 기간 표준/실적/생산성%
  [MM/DD] × N: 각 날짜의 구역별 표준/실적/생산성%/피킹금액/피킹박스/WMS시간

Usage
  python scripts/generate_report.py --start 2026-07-01 --end 2026-07-05
  python scripts/generate_report.py --dates 2026-07-01 2026-07-02 2026-07-03
  python scripts/generate_report.py --start 2026-07-01 --end 2026-07-05 --out 내보고서.xlsx
"""
import argparse
import json
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR  = Path(__file__).parent.parent
TEMP_DIR  = BASE_DIR / "data" / "temp"
DAILY_DIR = BASE_DIR / "data" / "daily"
OUT_DIR   = BASE_DIR / "data" / "reports"

ZONE_ORDER = ["H-I", "C-D", "A-P", "DPS", "E-F", "J-K", "L", "B", "L/S",
              "M-N", "S", "W", "R"]


# ── 데이터 로드 ─────────────────────────────────────────────────────────
def load_zones(d: date) -> list[dict] | None:
    """zones_{date}.json 로드. temp → daily 순서로 탐색"""
    candidates = [
        TEMP_DIR / f"zones_{d}.json",
        DAILY_DIR / d.strftime("%Y-%m") / f"zones_{d}.json",
    ]
    for p in candidates:
        if p.exists():
            return json.loads(p.read_text(encoding="utf-8"))
    return None


def load_all(dates: list[date]) -> dict[date, list[dict]]:
    """날짜 목록 → {date: rows} dict. 데이터 없는 날짜는 제외"""
    result = {}
    for d in dates:
        rows = load_zones(d)
        if rows:
            result[d] = rows
        else:
            print(f"  [SKIP] {d} — zones JSON 없음")
    return result


# ── 스타일 헬퍼 ─────────────────────────────────────────────────────────
_THIN = Side(style="thin")
_THICK = Side(style="medium")

def _border(top=False, bottom=False, left=False, right=False, thick_bottom=False):
    return Border(
        top    = _THIN  if top    else Side(style=None),
        bottom = _THICK if thick_bottom else (_THIN if bottom else Side(style=None)),
        left   = _THIN  if left   else Side(style=None),
        right  = _THIN  if right  else Side(style=None),
    )

def _fill(hex_color: str):
    return PatternFill("solid", fgColor=hex_color)

HDR_FILL   = _fill("1F4E79")   # 진한 파랑 (헤더)
SUB_FILL   = _fill("2F75B6")   # 중간 파랑 (서브헤더)
SUM_FILL   = _fill("D6E4F7")   # 연한 파랑 (합계행)
BAND_FILL  = _fill("F2F7FC")   # 교번 줄무늬
WHITE_FILL = _fill("FFFFFF")

HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
BOLD_FONT = Font(bold=True, size=10)
BASE_FONT = Font(size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")

def _hdr(ws, row, col, value, fill=None, font=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill  = fill  or HDR_FILL
    c.font  = font  or HDR_FONT
    c.alignment = align or CENTER
    return c

def _cell(ws, row, col, value=None, fmt=None, fill=None, font=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if fmt:    c.number_format = fmt
    if fill:   c.fill = fill
    c.font  = font  or BASE_FONT
    c.alignment = align or RIGHT
    if border: c.border = border
    return c


# ── 합계 시트 ───────────────────────────────────────────────────────────
def build_summary(wb: openpyxl.Workbook, data: dict[date, list[dict]]):
    ws = wb.create_sheet("합계", 0)
    dates = sorted(data.keys())

    # ── 컬럼 레이아웃 ──────────────────────────────────────────────────
    # 고정: [구역(1)] + [날짜별 실적hr(1col×N)] + [기간합계: 표준hr, 실적hr, 생산성%]
    COL_ZONE   = 1
    COL_DATES  = {d: 2 + i for i, d in enumerate(dates)}   # 날짜별 열
    COL_STD    = 2 + len(dates)       # 기간 표준
    COL_ACT    = COL_STD + 1          # 기간 실적
    COL_PROD   = COL_ACT  + 1         # 생산성%
    TOTAL_COLS = COL_PROD

    # ── 1행: 헤더 ──────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 30
    ws.column_dimensions["A"].width = 10

    _hdr(ws, 1, COL_ZONE, "구역")
    for d, col in COL_DATES.items():
        _hdr(ws, 1, col, f"{d.month}.{d.day}")
        ws.column_dimensions[get_column_letter(col)].width = 10
    _hdr(ws, 1, COL_STD,  "기간\n표준(hr)", fill=_fill("2F75B6"))
    _hdr(ws, 1, COL_ACT,  "기간\n실적(hr)", fill=_fill("2F75B6"))
    _hdr(ws, 1, COL_PROD, "생산성%",        fill=_fill("2F75B6"))
    for col in [COL_STD, COL_ACT, COL_PROD]:
        ws.column_dimensions[get_column_letter(col)].width = 11

    # ── 구역별 집계 ────────────────────────────────────────────────────
    # {zone: {date: {std, act}}}
    zone_data: dict[str, dict] = {z: {} for z in ZONE_ORDER}
    for d, rows in data.items():
        for r in rows:
            z = r["zone"]
            if z in zone_data:
                zone_data[z][d] = {
                    "std": r.get("std_time_hr", 0) or 0,
                    "act": r.get("act_time_hr", 0) or 0,
                }

    # ── 데이터 행 ──────────────────────────────────────────────────────
    # 실제 데이터가 있는 구역만 출력
    active_zones = [z for z in ZONE_ORDER if any(zone_data[z].values())]

    for i, zone in enumerate(active_zones):
        row = i + 2
        fill = BAND_FILL if i % 2 == 0 else WHITE_FILL

        _cell(ws, row, COL_ZONE, zone, fill=fill, font=BOLD_FONT, align=CENTER)

        period_std = period_act = 0.0
        for d, col in COL_DATES.items():
            v = zone_data[zone].get(d, {})
            act = v.get("act", 0)
            std = v.get("std", 0)
            _cell(ws, row, col, act if act else None, fmt='0.00', fill=fill)
            period_std += std
            period_act += act

        _cell(ws, row, COL_STD,  round(period_std, 4) if period_std else None, fmt='0.00', fill=fill)
        _cell(ws, row, COL_ACT,  round(period_act, 4) if period_act else None, fmt='0.00', fill=fill)
        prod = (period_std / period_act * 100) if period_act else None
        _cell(ws, row, COL_PROD, round(prod, 1) if prod else None,
              fmt='0.0"%"', fill=fill)

    # ── 합계 행 ────────────────────────────────────────────────────────
    sum_row = 2 + len(active_zones)
    ws.row_dimensions[sum_row].height = 18
    _cell(ws, sum_row, COL_ZONE, "합계", fill=SUM_FILL, font=BOLD_FONT, align=CENTER)

    total_std = total_act = 0.0
    for d, col in COL_DATES.items():
        col_total = sum(
            zone_data[z].get(d, {}).get("act", 0) for z in active_zones
        )
        _cell(ws, sum_row, col, round(col_total, 2) if col_total else None,
              fmt='0.00', fill=SUM_FILL, font=BOLD_FONT)

    for z in active_zones:
        for d in dates:
            total_std += zone_data[z].get(d, {}).get("std", 0)
            total_act += zone_data[z].get(d, {}).get("act", 0)

    _cell(ws, sum_row, COL_STD,  round(total_std, 4) if total_std else None,
          fmt='0.00', fill=SUM_FILL, font=BOLD_FONT)
    _cell(ws, sum_row, COL_ACT,  round(total_act, 4) if total_act else None,
          fmt='0.00', fill=SUM_FILL, font=BOLD_FONT)
    prod_total = (total_std / total_act * 100) if total_act else None
    _cell(ws, sum_row, COL_PROD, round(prod_total, 1) if prod_total else None,
          fmt='0.0"%"', fill=SUM_FILL, font=BOLD_FONT)

    # ── 전체 테두리 ────────────────────────────────────────────────────
    last_row = sum_row
    for row in range(1, last_row + 1):
        for col in range(1, TOTAL_COLS + 1):
            c = ws.cell(row=row, column=col)
            c.border = Border(
                top    = _THIN,
                bottom = _THICK if row == last_row else _THIN,
                left   = _THIN,
                right  = _THICK if col == TOTAL_COLS else _THIN,
            )

    # 열 고정 (날짜 컬럼 스크롤 시 구역명 고정)
    ws.freeze_panes = "B2"


# ── 개별 날짜 시트 ──────────────────────────────────────────────────────
def build_daily(wb: openpyxl.Workbook, d: date, rows: list[dict]):
    sheet_name = f"{d.month}.{d.day}"   # "/" 는 Excel 시트명 불가 → 점 사용
    ws = wb.create_sheet(sheet_name)

    # ── 컬럼 정의 ──────────────────────────────────────────────────────
    COLS = [
        ("구역",          12, None),
        ("표준시간(hr)",  13, '0.0000'),
        ("실적시간(hr)",  13, '0.0000'),
        ("생산성%",       10, '0.0"%"'),
        ("피킹금액",      16, '#,##0'),
        ("피킹박스",      10, '#,##0'),
        ("WMS근무(hr)",   12, '0.00'),
    ]
    COL_NAMES  = [c[0] for c in COLS]
    COL_WIDTHS = [c[1] for c in COLS]
    COL_FMTS   = [c[2] for c in COLS]
    N = len(COLS)

    # ── 1행: 날짜 헤더 (병합) ──────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N)
    title = ws.cell(row=1, column=1,
                    value=f"피킹 생산성  {d.year}년 {d.month}월 {d.day}일")
    title.fill  = _fill("1F4E79")
    title.font  = Font(bold=True, color="FFFFFF", size=12)
    title.alignment = CENTER
    ws.row_dimensions[1].height = 28

    # ── 2행: 컬럼 헤더 ─────────────────────────────────────────────────
    ws.row_dimensions[2].height = 20
    for j, (name, width, _) in enumerate(COLS, 1):
        _hdr(ws, 2, j, name, fill=_fill("2F75B6"))
        ws.column_dimensions[get_column_letter(j)].width = width

    # ── 행 집계 ────────────────────────────────────────────────────────
    zone_rows: dict[str, dict] = {}
    for r in rows:
        z = r["zone"]
        if z in ZONE_ORDER:
            zone_rows[z] = r

    active = [z for z in ZONE_ORDER if z in zone_rows]

    for i, zone in enumerate(active):
        row = i + 3
        r = zone_rows[zone]
        std = r.get("std_time_hr") or 0
        act = r.get("act_time_hr") or 0
        prod = (std / act * 100) if act else None
        fill = BAND_FILL if i % 2 == 0 else WHITE_FILL

        values = [
            zone,
            round(std, 4),
            round(act, 4),
            round(prod, 1) if prod else None,
            r.get("pick_amount") or None,
            r.get("pick_box") or None,
            round(r.get("wms_time_hr") or 0, 2) or None,
        ]
        for j, (val, fmt) in enumerate(zip(values, COL_FMTS), 1):
            align = CENTER if j == 1 else RIGHT
            _cell(ws, row, j, val, fmt=fmt, fill=fill,
                  font=BOLD_FONT if j == 1 else BASE_FONT, align=align)

    # ── 합계 행 ────────────────────────────────────────────────────────
    sum_row = 3 + len(active)
    ws.row_dimensions[sum_row].height = 18

    tot_std = sum((zone_rows[z].get("std_time_hr") or 0) for z in active)
    tot_act = sum((zone_rows[z].get("act_time_hr") or 0) for z in active)
    tot_amt = sum((zone_rows[z].get("pick_amount") or 0) for z in active)
    tot_box = sum((zone_rows[z].get("pick_box")    or 0) for z in active)
    tot_wms = sum((zone_rows[z].get("wms_time_hr") or 0) for z in active)
    tot_prd = (tot_std / tot_act * 100) if tot_act else None

    sum_vals = [
        "합계",
        round(tot_std, 4),
        round(tot_act, 4),
        round(tot_prd, 1) if tot_prd else None,
        round(tot_amt),
        int(tot_box),
        round(tot_wms, 2),
    ]
    for j, (val, fmt) in enumerate(zip(sum_vals, COL_FMTS), 1):
        align = CENTER if j == 1 else RIGHT
        _cell(ws, sum_row, j, val, fmt=fmt, fill=SUM_FILL, font=BOLD_FONT, align=align)

    # ── 테두리 ─────────────────────────────────────────────────────────
    for row in range(2, sum_row + 1):
        for col in range(1, N + 1):
            ws.cell(row=row, column=col).border = Border(
                top    = _THIN,
                bottom = _THICK if row == sum_row else _THIN,
                left   = _THIN,
                right  = _THICK if col == N else _THIN,
            )


# ── 메인 ──────────────────────────────────────────────────────────────
def date_range(start: date, end: date):
    d = start
    while d <= end:
        yield d
        d += timedelta(days=1)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--start", help="시작 날짜 YYYY-MM-DD")
    ap.add_argument("--end",   help="종료 날짜 YYYY-MM-DD")
    ap.add_argument("--dates", nargs="+", help="특정 날짜 목록 (--start/--end 대신)")
    ap.add_argument("--out",   help="출력 파일명 (기본: 피킹생산성_MMDD_MMDD.xlsx)")
    args = ap.parse_args()

    if args.dates:
        dates = sorted(datetime.strptime(d, "%Y-%m-%d").date() for d in args.dates)
    elif args.start and args.end:
        start = datetime.strptime(args.start, "%Y-%m-%d").date()
        end   = datetime.strptime(args.end,   "%Y-%m-%d").date()
        dates = list(date_range(start, end))
    else:
        ap.error("--start/--end 또는 --dates 필요")

    print(f"\n리포트 생성 ({len(dates)}개 날짜)...")
    data = load_all(dates)

    if not data:
        print("데이터 없음. zones_{date}.json 파일 확인 필요.")
        return

    loaded_dates = sorted(data.keys())
    print(f"  로드 완료: {loaded_dates[0]} ~ {loaded_dates[-1]} ({len(data)}일)")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # 기본 빈 시트 제거

    build_summary(wb, data)
    for d in loaded_dates:
        build_daily(wb, d, data[d])

    # 출력 경로
    if args.out:
        out_path = Path(args.out)
        if not out_path.is_absolute():
            out_path = BASE_DIR / args.out
    else:
        s = loaded_dates[0].strftime("%m%d")
        e = loaded_dates[-1].strftime("%m%d")
        fname = f"피킹생산성_{s}_{e}.xlsx" if s != e else f"피킹생산성_{s}.xlsx"
        out_path = OUT_DIR / fname

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))

    print(f"\n저장 완료: {out_path}")
    print(f"  시트: 합계 + {len(loaded_dates)}개 일자 ({', '.join(str(d.month)+'/'+str(d.day) for d in loaded_dates)})")


if __name__ == "__main__":
    main()
