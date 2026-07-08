# -*- coding: utf-8 -*-
"""
입고 생산성 자동화 (일룸/데스커) — v2

raw 2종(입고/이동) → 일자별 작업자 실적 집계 (일반 + 정산용 이중 분류)
→ JSON 저장 + 실적파일 2종(일반 '종합' 체계 / 정산용 '입고_종합' 체계) 대조 검증

브랜드별 차이:
  일룸   이동필터 QCC·IPCBS-00·REC-04 / 검사이동 To조건 {Y-IPCBS-00, Y-QCC-00}
  데스커 이동필터 QCC·REC-04         / 검사이동 To조건 {Y-REC-01, Y-QCC-00}
  퍼시스 (미구현 — 정상/반품 2유형 단순 구조, raw 확보 후 추가)

입고유형 2종 (같은 행에 동시 계산):
  [일반 J]   To=Y-REC-00→CUT / raw=반품입고→반품입고 / 나머지→정상입고
  [정산용 J] From=Y-REC-04→정품화입고
             From=Y-QCC-00→재입고
             To∈검사이동조건→"검사이동,업체반송"
             To=Y-REC-00→CUT
             raw=반품입고→반품입고 / 나머지→정상입고
  ※ 정산용 입고_종합은 CUT 제외 5유형 집계 (CUT은 별도 cut 시트 — 별도 raw, 미구현)

공통 로직 (v1에서 검증 완료):
  - MOVER_RFxxx → MOVER / 작업자명 raw 그대로(태그 유지)
  - 금액 = data/master/기준정보_공장도가.xlsx (브랜드별 단품정보 시트) D열→F열 × 수량
    (2026-07-07: 브랜드별 마스터 자체 시트 → 통합 파일로 전환, 4브랜드 전부 이 파일 사용)
  - 파렛트 = CUT 제외 행당 1 (일반/정산용 J 각각 기준)
  - 소요시간(min) = 같은 작업자 연속 스캔 간격, 12→13시 -50, 17→18시 -30

Usage:
  python scripts/inbound_automation.py --date 2026-07-01 --brand 일룸
  python scripts/inbound_automation.py --date 2026-07-01 --brand 데스커 --no-verify
"""
import argparse
import json
import re
from datetime import date, datetime
from pathlib import Path

import pandas as pd

BASE_DIR   = Path(__file__).resolve().parent.parent
RAW_BASE   = BASE_DIR / "data" / "raw" / "2026_입고"   # 하위 월 폴더(07 등)에 raw
MASTER_DIR = BASE_DIR / "data" / "master"               # 실적 마스터 5종
TEMP_DIR   = BASE_DIR / "data" / "temp"


def raw_dir(target: date) -> Path:
    return RAW_BASE / target.strftime("%m")

BRANDS = {
    "일룸": {
        "file_basic":   "26년 07월 입고 실적_일룸.xlsx",
        "file_detail":  "26년 07월 입고 실적_일룸_IPC 정산용_최종버전.xlsx",
        "move_locs":    {"Y-QCC-00", "Y-IPCBS-00", "Y-REC-04"},
        "inspect_locs": {"Y-IPCBS-00", "Y-QCC-00"},
        # 공장도가는 통합 마스터에서 조회 (브랜드별 마스터 자체 시트와 100% 동일 — 코드/단가
        # 갱신 시 한 곳(기준정보_공장도가.xlsx)만 관리하도록 전환, 2026-07-07 재검증)
        "price_file":   "기준정보_공장도가.xlsx",
        "price_sheet":  "일-단품정보",
    },
    "데스커": {
        "file_basic":   "26년 07월 입고 실적_데스커.xlsx",
        "file_detail":  "26년 07월 입고 실적_데스커_승현 정산용_최종버전.xlsx",
        "move_locs":    {"Y-QCC-00", "Y-REC-04"},
        "inspect_locs": {"Y-REC-01", "Y-QCC-00"},
        "price_file":   "기준정보_공장도가.xlsx",
        "price_sheet":  "일-단품정보",
    },
    "퍼시스": {
        "file_basic":   "26년 07월 입고 실적_퍼시스.xlsx",
        "file_detail":  None,                       # 정산용 파일 없음 (일반 3유형만)
        # 퍼시스는 From 방향만 포함: 로케이션→Y-OUT-00(출고작업)은 입고실적 아님
        # (7/3~7/4 검증으로 확인. From=Y-OUT-00/QCC 실사례는 아직 없음 — 나오면 재검증)
        "move_from_only": True,
        "move_locs":    {"Y-QCC-00", "Y-OUT-00"},
        "inspect_locs": set(),                      # detail 분류 미사용
        "price_file":   "기준정보_공장도가.xlsx",
        "price_sheet":  "퍼,시_단품정보",
    },
    "3PL": {
        # 실적파일(마스터) 자체가 아직 없음 — 구조를 처음 설계하는 단계라 검증 없이 진행,
        # 추후 담당자 피드백으로 로직 보정 예정
        "file_basic":   None,
        "file_detail":  None,
        # raw에 여러 owner(client 브랜드)가 섞여 있고 그룹사(일룸/퍼시스/데스커 등) 물량도
        # 혼입되어 있음 — 그룹사는 각자 전용 파이프라인이 있으므로 3PL 집계에서 제외
        "exclude_owners": {"퍼시스", "일룸", "데스커", "시디즈", "알로소", "슬로우베드", "팀스"},
        # 3PL 전용 임시 스테이징 로케이션(Y-TCT-xx, 슬롯 번호가 여러 개) — QCC와 동일 역할
        "move_prefix":  "Y-TCT",
        "inspect_locs": set(),
        # 3PL 전용 단품정보 없음 → 통합 공장도가 마스터('바-단품정보') 사용
        "price_file":   "기준정보_공장도가.xlsx",
        "price_sheet":  "바-단품정보",
    },
}

TYPES_BASIC  = ["정상입고", "반품입고", "CUT"]
TYPES_DETAIL = ["정상입고", "반품입고", "정품화입고", "재입고", "검사이동,업체반송", "CUT"]

_TAG_RE = re.compile(r"^\[(주간|야간)\]")


def strip_tag(name: str) -> str:
    return _TAG_RE.sub("", str(name)).strip()


def norm_worker(name: str) -> str:
    s = str(name).strip()
    if s.upper().startswith("MOVER_RF"):
        return "MOVER"
    return s


# ── raw 로드 ──────────────────────────────────────────────────────────
def load_inbound(brand: str, target: date) -> pd.DataFrame:
    mmdd = target.strftime("%m%d")
    path = raw_dir(target) / f"입고_{brand}_{mmdd}.xlsx"
    if not path.exists():
        raise FileNotFoundError(path)
    df = pd.read_excel(path)
    df.columns = df.columns.str.strip()

    exclude = BRANDS[brand].get("exclude_owners")
    if exclude:
        before = len(df)
        df = df[~df["OWNER"].astype(str).str.strip().isin(exclude)]
        print(f"  입고 raw: 그룹사 {before - len(df)}행 제외")

    out = pd.DataFrame({
        "work_dt":  pd.to_datetime(df["작업일시"], format="mixed"),
        "owner":    df["OWNER"].astype(str).str.strip(),
        "worker":   df["작업자"].astype(str).map(norm_worker),
        "item":     df["ITEM ID"].astype(str).str.strip(),
        "qty":      df["수량"].astype(float),
        "from_loc": "",
        "to_loc":   df["LOCATION"].astype(str).str.strip(),
        "raw_type": df["입고 유형"].astype(str).str.strip(),
        "src":      "입고",
    })
    print(f"  입고 raw: {len(out)}행 ({path.name})")
    return out


def load_moves(brand: str, target: date) -> pd.DataFrame:
    mmdd = target.strftime("%m%d")
    path = raw_dir(target) / f"이동_{brand}_{mmdd}.xlsx"
    if not path.exists():
        raise FileNotFoundError(path)
    df = pd.read_excel(path)
    df.columns = df.columns.str.strip()

    exclude = BRANDS[brand].get("exclude_owners")
    if exclude:
        before = len(df)
        df = df[~df["OWNER"].astype(str).str.strip().isin(exclude)]
        print(f"  이동 raw: 그룹사 {before - len(df)}행 제외")

    frm = df["From Location"].astype(str).str.strip()
    to  = df["To Location"].astype(str).str.strip()

    move_prefix = BRANDS[brand].get("move_prefix")
    if move_prefix:
        mask = frm.str.startswith(move_prefix) | to.str.startswith(move_prefix)
    else:
        move_locs = BRANDS[brand]["move_locs"]
        if BRANDS[brand].get("move_from_only"):
            mask = frm.isin(move_locs)
        else:
            mask = frm.isin(move_locs) | to.isin(move_locs)
    sel = df[mask]

    out = pd.DataFrame({
        "work_dt":  pd.to_datetime(sel["작업일시"], format="mixed"),
        "owner":    sel["OWNER"].astype(str).str.strip(),
        "worker":   sel["작업자"].astype(str).map(norm_worker),
        "item":     sel["ITEM ID"].astype(str).str.strip(),
        "qty":      sel["이동수량"].astype(float),
        "from_loc": frm[mask],
        "to_loc":   to[mask],
        "raw_type": "",
        "src":      "이동",
    })
    print(f"  이동 raw: {len(df)}행 → 필터 후 {len(out)}행 ({path.name})")
    return out


def load_price_map(brand: str) -> dict:
    price_file = BRANDS[brand].get("price_file") or BRANDS[brand]["file_basic"]
    master = MASTER_DIR / price_file
    df = pd.read_excel(master, sheet_name=BRANDS[brand]["price_sheet"], usecols="D,F", header=0)
    df.columns = ["item", "price"]
    df = df.dropna(subset=["item"])
    df["item"] = df["item"].astype(str).str.strip()
    # 중복 품목은 첫 행 우선 — Excel VLOOKUP과 동일 (퍼시스: 가격 다른 중복 2,710종 존재)
    df = df.drop_duplicates(subset="item", keep="first")
    m = dict(zip(df["item"], df["price"]))
    print(f"  단품정보: {len(m)}건 ({price_file})")
    return m


# ── 계산 ─────────────────────────────────────────────────────────────
def build_daily(brand: str, target: date) -> pd.DataFrame:
    df = pd.concat([load_inbound(brand, target), load_moves(brand, target)],
                   ignore_index=True)
    df = df.sort_values(["worker", "work_dt"], kind="stable").reset_index(drop=True)

    inspect_locs = BRANDS[brand]["inspect_locs"]

    def classify_basic(r):
        if r["to_loc"] == "Y-REC-00":
            return "CUT"
        if r["raw_type"] == "반품입고":
            return "반품입고"
        return "정상입고"

    def classify_detail(r):
        if r["from_loc"] == "Y-REC-04":
            return "정품화입고"
        if r["from_loc"] == "Y-QCC-00":
            return "재입고"
        if r["to_loc"] in inspect_locs:
            return "검사이동,업체반송"
        if r["to_loc"] == "Y-REC-00":
            return "CUT"
        if r["raw_type"] == "반품입고":
            return "반품입고"
        return "정상입고"

    df["type_basic"]  = df.apply(classify_basic, axis=1)
    df["type_detail"] = df.apply(classify_detail, axis=1)

    price = load_price_map(brand)
    df["price"] = df["item"].map(price)
    missing_mask = df["price"].isna()
    missing = df[missing_mask]["item"].unique()
    gap_path = TEMP_DIR / f"price_gaps_inbound_{brand}_{target}.json"
    if len(missing):
        print(f"  [경고] 단가 없는 품목 {len(missing)}종 → 금액 0: {list(missing)[:5]}{' ...' if len(missing) > 5 else ''}")
        gaps = {}
        for item, g in df[missing_mask].groupby("item"):
            gaps[item] = {"qty": float(g["qty"].sum()), "rows": int(len(g))}
        gap_path.write_text(json.dumps(gaps, ensure_ascii=False, indent=2), encoding="utf-8")
    elif gap_path.exists():
        gap_path.unlink()  # 이전 실행의 stale 갭 파일 제거
    df["amount"] = (df["price"].fillna(0) * df["qty"]).round(0)

    df["pallet_basic"]  = (df["type_basic"]  != "CUT").astype(int)
    df["pallet_detail"] = (df["type_detail"] != "CUT").astype(int)

    # 소요시간(min)
    t = df["work_dt"].dt.round("s")
    mins = []
    for i in range(len(df)):
        if i == 0 or df.at[i, "worker"] != df.at[i - 1, "worker"]:
            mins.append(0.0)
            continue
        gap = (t.iloc[i] - t.iloc[i - 1]).total_seconds() / 60.0
        h_prev, h_cur = t.iloc[i - 1].hour, t.iloc[i].hour
        if h_prev == 12 and h_cur == 13:
            gap -= 50
        elif h_prev == 17 and h_cur == 18:
            gap -= 30
        mins.append(gap)
    df["dur_min"] = mins
    return df


_TYPE_KEYS = {
    "정상입고": "normal", "반품입고": "return", "정품화입고": "certify",
    "재입고": "reentry", "검사이동,업체반송": "inspect", "CUT": "cut",
}


def aggregate(df: pd.DataFrame, brand: str, target: date) -> list[dict]:
    rows = []
    for worker, g in df.groupby("worker", sort=True):
        rec = {
            "work_date": str(target),
            "brand": brand,
            "worker": worker,
            "worker_display": strip_tag(worker),
            "rows": len(g),
            "hours": round(g["dur_min"].sum() / 60.0, 6),
        }
        for scheme, tcol, pcol in [("basic", "type_basic", "pallet_basic"),
                                   ("detail", "type_detail", "pallet_detail")]:
            types = TYPES_BASIC if scheme == "basic" else TYPES_DETAIL
            for tp in types:
                sub = g[g[tcol] == tp]
                key = _TYPE_KEYS[tp]
                rec[f"{scheme}_qty_{key}"] = float(sub["qty"].sum())
                rec[f"{scheme}_amt_{key}"] = float(sub["amount"].sum())
            rec[f"{scheme}_pallets"] = int(g[pcol].sum())
        rec["qty_total"] = float(g["qty"].sum())
        rec["amt_total"] = float(g["amount"].sum())
        rows.append(rec)
    return rows


# ── 검증 ─────────────────────────────────────────────────────────────
def _read_sheet_agg(path: Path, sheet: str) -> dict | None:
    """실적파일 일자시트 → {작업자: {qty:{유형:합}, amt:{...}, pallet, min, rows}}"""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return None
    ws = wb[sheet]
    actual: dict[str, dict] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        w = str(r[2])
        a = actual.setdefault(w, {"qty": {}, "amt": {}, "pallet": 0, "min": 0.0, "rows": 0})
        j = str(r[9]) if r[9] is not None else ""
        a["qty"][j] = a["qty"].get(j, 0) + (r[11] or 0)
        a["amt"][j] = a["amt"].get(j, 0) + (r[12] or 0)
        a["pallet"] += (r[13] or 0)
        a["min"]    += (r[19] or 0)
        a["rows"]   += 1
    wb.close()
    return actual


def _compare(label: str, actual: dict, df: pd.DataFrame,
             tcol: str, pcol: str, types: list[str]) -> bool:
    mine: dict[str, dict] = {}
    for w, g in df.groupby(df["worker"].map(strip_tag)):
        mine[w] = {
            "qty": g.groupby(tcol)["qty"].sum().to_dict(),
            "amt": g.groupby(tcol)["amount"].sum().to_dict(),
            "pallet": int(g[pcol].sum()),
            "min": float(g["dur_min"].sum()),
            "rows": len(g),
        }
    print(f"\n[검증-{label}]")
    ok = True
    for w in sorted(set(actual) | set(mine)):
        a, m = actual.get(w), mine.get(w)
        if a is None:
            print(f"  [차이] {w}: 시트에 없음 (자동화 {m['rows']}행)"); ok = False; continue
        if m is None:
            print(f"  [차이] {w}: 자동화에 없음 (시트 {a['rows']}행)"); ok = False; continue
        diffs = []
        if a["rows"] != m["rows"]:
            diffs.append(f"행수 {a['rows']}≠{m['rows']}")
        for tp in types:
            aq, mq = a["qty"].get(tp, 0), m["qty"].get(tp, 0)
            if abs(aq - mq) > 0.01:
                diffs.append(f"{tp}수량 {aq}≠{mq}")
            aa, ma = a["amt"].get(tp, 0), m["amt"].get(tp, 0)
            if abs(aa - ma) > 1:
                diffs.append(f"{tp}금액 {aa:,.0f}≠{ma:,.0f}")
        if a["pallet"] != m["pallet"]:
            diffs.append(f"파렛트 {a['pallet']}≠{m['pallet']}")
        if abs(a["min"] - m["min"]) > 0.5:
            diffs.append(f"소요시간 {a['min']:.2f}≠{m['min']:.2f}")
        if diffs:
            ok = False
            print(f"  [차이] {w}: " + " / ".join(diffs))
        else:
            print(f"  [일치] {w}: {m['rows']}행, {m['min']:.1f}min")
    return ok


def verify(df: pd.DataFrame, brand: str, target: date):
    sheet = target.strftime("%m.%d")
    cfg = BRANDS[brand]
    results = []
    for label, fkey, tcol, pcol, types in [
        ("일반",   "file_basic",  "type_basic",  "pallet_basic",  TYPES_BASIC),
        ("정산용", "file_detail", "type_detail", "pallet_detail", TYPES_DETAIL),
    ]:
        if cfg[fkey] is None:
            continue
        actual = _read_sheet_agg(MASTER_DIR / cfg[fkey], sheet)
        if actual is None:
            print(f"\n[검증-{label} 생략] '{sheet}' 시트 없음")
            continue
        results.append(_compare(f"{label}({cfg[fkey][:20]}…)", actual, df, tcol, pcol, types))
    if results:
        print(f"\n검증 결과: {'전체 일치' if all(results) else '차이 있음 — 위 항목 확인'}")
    elif cfg["file_basic"] is None:
        print(f"\n[검증 불가] {brand}: 실적 마스터 파일 없음 — 구조 신설 단계, 대조 없이 진행")


# ── 메인 ─────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--date",  required=True, help="YYYY-MM-DD")
    ap.add_argument("--brand", default="일룸", choices=list(BRANDS))
    ap.add_argument("--no-verify", action="store_true")
    args = ap.parse_args()
    target = datetime.strptime(args.date, "%Y-%m-%d").date()

    print(f"=== 입고 자동화 {args.brand} {target} ===")
    df = build_daily(args.brand, target)
    print(f"  병합 후: {len(df)}행, 작업자 {df['worker'].nunique()}명")

    rows = aggregate(df, args.brand, target)
    out = TEMP_DIR / f"inbound_{args.brand}_{target}.json"
    out.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  저장: {out}")

    tot_qty = sum(r["qty_total"] for r in rows)
    tot_amt = sum(r["amt_total"] for r in rows)
    tot_hr  = sum(r["hours"] for r in rows)
    print(f"  합계: 수량 {tot_qty:,.0f} / 금액 {tot_amt:,.0f} / 시간 {tot_hr:.2f}h")

    if not args.no_verify:
        verify(df, args.brand, target)


if __name__ == "__main__":
    main()
