# -*- coding: utf-8 -*-
"""
과거 월별 종합실적 Excel → picking_zone_daily DB 적재 (2023-01 ~ 2026-05)

사용법:
  python scripts/load_historical_zone.py --dry-run          # 파싱 검증
  python scripts/load_historical_zone.py --year 2023 --dry-run
  python scripts/load_historical_zone.py                    # 전체 적재

파일 위치: data/raw/2026/{연도년}/{파일명}.xlsx
"""
import argparse
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
import psycopg2
from psycopg2.extras import execute_values

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data/raw/2026"
load_dotenv(BASE_DIR / ".env")

# 구역 레이블 → (owner, center, DB 저장 구역명)
ZONE_MAP = {
    "H-I":  ("일룸",   "양지1센터", "H-I"),
    "C-D":  ("일룸",   "양지1센터", "C-D"),
    "A-P":  ("일룸",   "양지1센터", "A-P"),
    "A-B":  ("일룸",   "양지1센터", "A-P"),   # 구 이름 → A-P로 통일
    "DPS":  ("일룸",   "양지1센터", "DPS"),
    "E-F":  ("퍼시스", "양지1센터", "E-F"),
    "J-K":  ("퍼시스", "양지1센터", "J-K"),
    "L":    ("퍼시스", "양지1센터", "L"),
    "B":    ("퍼시스", "양지1센터", "B"),
    "L/S":  ("퍼시스", "양지1센터", "L/S"),
    "M-N":  ("데스커", "양지2센터", "M-N"),
    "S":    ("데스커", "양지2센터", "S"),
    "W":    ("3PL",    "양지3센터", "W"),
    "R":    ("3PL",    "양지3센터", "R"),
    "P":    ("3PL",    "양지3센터", "P"),
    "p":    ("3PL",    "양지3센터", "P"),   # 소문자 P (2023-03~2024-10)
}
SKIP_ZONES = {"H-300", "H300"}

DATE_ROW = 5    # 날짜가 있는 행 (1-based)
MAX_ROW  = 700
MAX_COL  = 35

# 총 피킹금액 행 기준 오프셋
OFF_BOX = 2   # 총 피킹 박스수
OFF_STD = 3   # 표준시간(작업시간)[hr]
OFF_ACT = 14  # 투입시간[hr]  (표준시간 행 +11 = 피킹금액 행 +14)


def _to_float(v) -> float:
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def _parse_date(v, year_hint: int):
    """datetime 객체 또는 텍스트 "M/D" 형식을 date로 변환"""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str) and "/" in v:
        parts = v.strip().split("/")
        if len(parts) == 2:
            try:
                return date(year_hint, int(parts[0]), int(parts[1]))
            except ValueError:
                pass
    return None


def _find_sheet(wb):
    for name in wb.sheetnames:
        if "종합실적" in name and "연도" not in name and "연간" not in name:
            return wb[name]
    return None


def parse_monthly_file(filepath: Path, year_hint: int) -> list:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = _find_sheet(wb)
    if ws is None:
        wb.close()
        return []

    # read_only 모드: 한 번에 전체 로드
    all_rows = {}
    for ri, row in enumerate(
        ws.iter_rows(min_row=1, max_row=MAX_ROW, max_col=MAX_COL, values_only=True),
        start=1,
    ):
        all_rows[ri] = list(row)
    wb.close()

    # 날짜 컬럼 수집 (5행)
    date_cols = {}  # col_idx(1-based) → date
    for ci, v in enumerate(all_rows.get(DATE_ROW, []), start=1):
        d = _parse_date(v, year_hint)
        if d:
            date_cols[ci] = d

    if not date_cols:
        return []

    # 구역 행 탐지: C열에 '피킹금액' 포함 → B열 우선, 없으면 A열
    zone_rows = []  # [(zone_label, row_idx)]
    for ri in sorted(all_rows):
        row_vals = all_rows[ri]
        c_val = row_vals[2] if len(row_vals) > 2 else None
        if c_val is None or "피킹금액" not in str(c_val):
            continue

        a_raw = row_vals[0]
        b_raw = row_vals[1]
        b_str = str(b_raw).strip() if b_raw is not None else ""
        a_str = str(a_raw).strip() if a_raw is not None else ""

        zone_label = b_str if b_str else a_str
        if not zone_label:
            continue
        if zone_label in SKIP_ZONES:
            continue
        if zone_label in ZONE_MAP:
            zone_rows.append((zone_label, ri))

    # 구역 × 날짜 레코드 생성
    records = []
    for zone_label, amount_row in zone_rows:
        owner, center, db_zone = ZONE_MAP[zone_label]
        box_row = amount_row + OFF_BOX
        std_row = amount_row + OFF_STD
        act_row = amount_row + OFF_ACT

        for col_idx, work_date in date_cols.items():
            ci0 = col_idx - 1  # 0-based

            def _get(row_idx, _ci=ci0):
                r = all_rows.get(row_idx, [])
                return r[_ci] if _ci < len(r) else None

            amount = _to_float(_get(amount_row))
            box    = _to_float(_get(box_row))
            std    = _to_float(_get(std_row))
            act    = _to_float(_get(act_row))

            if amount == 0 and box == 0 and std == 0 and act == 0:
                continue

            records.append({
                "work_date":   work_date,
                "center":      center,
                "owner":       owner,
                "zone":        db_zone,
                "std_time_hr": round(std, 4),
                "act_time_hr": round(act, 4),
                "pick_amount": int(amount),
                "pick_box":    int(box),
            })

    return records


def get_monthly_files(year_filter=None, month_filter=None):
    files = []
    for year_dir in sorted(DATA_DIR.glob("[0-9][0-9][0-9][0-9]년")):
        for f in sorted(year_dir.glob("*.xlsx")):
            m = re.search(r"(\d{4})년\s*(\d{1,2})월", f.stem)
            if not m:
                continue
            y, mo = int(m.group(1)), int(m.group(2))
            if year_filter and y != year_filter:
                continue
            if month_filter and mo != month_filter:
                continue
            files.append((y, mo, f))
    return sorted(files)


def db_connect():
    url = os.getenv("SUPABASE_POOLER_URL") or os.getenv("SUPABASE_DB_URL")
    if not url:
        sys.exit("DB URL 미설정 (SUPABASE_POOLER_URL)")
    return psycopg2.connect(url)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="파싱 검증만, DB 미적재")
    ap.add_argument("--year",  type=int, help="특정 연도만")
    ap.add_argument("--month", type=int, help="특정 월만 (--year 함께 사용)")
    args = ap.parse_args()

    files = get_monthly_files(args.year, args.month)
    if not files:
        print("대상 파일 없음")
        return

    print(f"처리 대상: {len(files)}개 파일\n")

    all_records = []
    for year, month, fpath in files:
        recs = parse_monthly_file(fpath, year)
        print(f"  {year}년 {month:02d}월  {fpath.name}  → {len(recs):4d}건")
        all_records.extend(recs)

    print(f"\n총 파싱: {len(all_records):,}건")

    if args.dry_run:
        print("\n[DRY RUN] DB 미적재 - 샘플 10건:")
        for r in all_records[:10]:
            print(
                f"  {r['work_date']}  {r['owner']:6s} {r['zone']:5s}"
                f"  금액={r['pick_amount']:>12,}  박스={r['pick_box']:>6,}"
                f"  std={r['std_time_hr']:.2f}h  act={r['act_time_hr']:.2f}h"
            )
        # 구역 통계
        from collections import Counter
        zone_cnt = Counter(r["zone"] for r in all_records)
        print("\n구역별 레코드 수:")
        for z, cnt in sorted(zone_cnt.items()):
            print(f"  {z:6s}: {cnt:,}건")
        return

    conn = db_connect()
    cur = conn.cursor()
    try:
        dates = sorted(set(r["work_date"] for r in all_records))
        print(f"날짜 범위: {dates[0]} ~ {dates[-1]}  ({len(dates)}일)")

        # 해당 날짜의 기존 데이터 삭제 (멱등)
        print("기존 데이터 삭제 중...")
        for d in dates:
            cur.execute("DELETE FROM picking_zone_daily WHERE work_date = %s", (d,))

        rows = [
            (r["work_date"], r["center"], r["owner"], r["zone"],
             r["std_time_hr"], r["act_time_hr"], r["pick_amount"], r["pick_box"])
            for r in all_records
        ]
        execute_values(
            cur,
            """
            INSERT INTO picking_zone_daily
                (work_date, center, owner, zone,
                 std_time_hr, act_time_hr, pick_amount, pick_box)
            VALUES %s
            """,
            rows,
        )
        conn.commit()
        print(f"[COMMIT] {len(rows):,}건 적재 완료\n")

        cur.execute("""
            SELECT DATE_TRUNC('month', work_date) AS ym,
                   COUNT(*)             AS rows,
                   COUNT(DISTINCT zone) AS zones
            FROM picking_zone_daily
            GROUP BY ym ORDER BY ym
        """)
        print("DB 현황 (전체):")
        for row in cur.fetchall():
            print(f"  {str(row[0])[:7]}  {row[1]:4d}건  {row[2]}개 구역")

    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    main()
