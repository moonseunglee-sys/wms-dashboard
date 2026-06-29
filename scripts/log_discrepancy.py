# -*- coding: utf-8 -*-
"""
검증 불일치 누적 로거.

zones_<date>.json(우리 계산값) vs 종합실적_최종(담당자 기준) 비교 →
0이 아닌 차이를 분류와 함께 data/discrepancy_log.csv 에 누적.
(나중에 한번에 원인 찾아 수정하기 위한 기록. 같은 날짜 재실행 시 해당 날짜분 갱신)

분류:
  DPS_self_ref        : DPS 표준/실적 (외부 API값, self-reference)
  group_order_nuance  : 비-DPS 표준 잔차 (그룹내 방문순서 차이, 박스는 일치)
  price_version_diff  : 금액 차이 (공장도가 단가 버전차, 박스 일치)
  box_mismatch        : 박스수 차이 (★진짜 조사 필요)
  act_diff            : 실적시간 차이 (그 외)
  investigate         : 위에 안 맞는 표준 차이 (★조사 필요)

사용: python scripts/log_discrepancy.py --date 2026-06-08
"""
import argparse, csv, json
from datetime import datetime, date as _date
from pathlib import Path
import openpyxl

BASE = Path(__file__).resolve().parent.parent
FINAL = BASE / "data/master/2026년 06월 피킹 종합실적_최종.xlsx"
LOG = BASE / "data/discrepancy_log.csv"
EPS = 0.05  # % 이하는 0(반올림오차)로 간주

ZONE_STD_ROW_FINAL = {
    "H-I":10,"C-D":34,"A-P":58,"DPS":106,"E-F":190,"J-K":214,
    "L":238,"B":262,"L/S":286,"M-N":355,"S":379,"W":519,"R":543,
}
HEADER = ["logged_at","work_date","owner","zone","metric","our","expected","pct","category"]


def _expected(date_str):
    y,m,d = map(int,date_str.split("-"))
    col = 4 + (_date(y,m,d) - _date(2026,6,1)).days
    wb = openpyxl.load_workbook(FINAL, read_only=True, data_only=True); ws = wb.worksheets[2]
    out = {}
    for z,r in ZONE_STD_ROW_FINAL.items():
        std = ws.cell(r,col).value
        if isinstance(std,(int,float)) and std>0:
            out[z] = {"std":float(std), "amount":float(ws.cell(r-3,col).value or 0),
                      "box":float(ws.cell(r-1,col).value or 0), "act":float(ws.cell(r+11,col).value or 0)}
    wb.close(); return out


def _category(metric, zone, box_ok):
    if metric == "box":      return "box_mismatch"
    if metric == "amount":   return "price_version_diff" if box_ok else "investigate"
    if metric == "std":
        if zone == "DPS":    return "DPS_self_ref"
        return "group_order_nuance" if box_ok else "investigate"
    if metric == "act":
        if zone == "DPS":    return "DPS_self_ref"
        return "act_diff"
    return "investigate"


def main():
    ap = argparse.ArgumentParser(); ap.add_argument("--date", required=True); args = ap.parse_args()
    ds = args.date
    zpath = BASE / f"data/temp/zones_{ds}.json"
    if not zpath.exists():
        print(f"  [skip] {zpath.name} 없음 — 자동화 먼저 실행"); return
    ours = {d["zone"]: d for d in json.loads(zpath.read_text(encoding="utf-8"))}
    exp = _expected(ds)
    now = datetime.now().isoformat(timespec="seconds")

    # 기존 로그 로드 (해당 날짜 제외하고 보존 → 멱등)
    existing = []
    if LOG.exists():
        with open(LOG, encoding="utf-8-sig") as f:
            existing = [r for r in csv.DictReader(f) if r["work_date"] != ds]

    new_rows = []
    for zone, e in exp.items():
        o = ours.get(zone, {})
        owner = o.get("owner","")
        box_ok = abs((o.get("pick_box") or 0) - e["box"]) <= max(1, e["box"]*EPS/100)
        checks = [
            ("std",    o.get("std_time_hr") or 0, e["std"]),
            ("act",    o.get("act_time_hr") or 0, e["act"]),
            ("box",    o.get("pick_box") or 0,    e["box"]),
            ("amount", o.get("pick_amount") or 0, e["amount"]),
        ]
        for metric, ov, ev in checks:
            if not ev: continue
            pct = abs(ov-ev)/ev*100
            if pct <= EPS: continue
            new_rows.append({
                "logged_at":now, "work_date":ds, "owner":owner, "zone":zone, "metric":metric,
                "our":round(ov,4), "expected":round(ev,4), "pct":round(pct,3),
                "category":_category(metric, zone, box_ok),
            })

    with open(LOG, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=HEADER); w.writeheader()
        w.writerows(existing + new_rows)

    print(f"  [{ds}] 불일치 {len(new_rows)}건 기록 → {LOG.name}")
    from collections import Counter
    c = Counter(r["category"] for r in new_rows)
    for cat,n in c.most_common(): print(f"     {cat}: {n}")
    inv = [r for r in new_rows if r["category"] in ("investigate","box_mismatch")]
    if inv:
        detail = ", ".join("{}/{} {}%".format(r["zone"], r["metric"], r["pct"]) for r in inv)
        print("     ★조사필요:", detail)


if __name__ == "__main__":
    main()
