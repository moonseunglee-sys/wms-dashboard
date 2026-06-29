"""
picking_automation_v2.py  –  피킹 생산성 자동화 (v3)

Usage:
  python scripts/picking_automation_v2.py --date 2026-05-12

변경 이력:
  v3: F_1/D_1/DU_1 V열(22) 처리 추가, 정렬 WAVE번호(H) 기준으로 통일
"""

import argparse
import glob
import os
import re
import shutil
import subprocess
import time
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
import psycopg2
from dotenv import load_dotenv
from psycopg2.extras import execute_values

# ── 경로 ─────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent.parent
load_dotenv(BASE_DIR / ".env")

MASTER1 = BASE_DIR / "data/master/기준정보_마스터.xlsx"
MASTER2 = BASE_DIR / "data/master/기준정보2_마스터.xlsx"
TEMP1   = BASE_DIR / "data/temp/tmp_master1.xlsx"
TEMP2   = BASE_DIR / "data/temp/tmp_master2.xlsx"
RAW_DIR = BASE_DIR / "data/raw"

KGA1_PATH = BASE_DIR / "data/master/가동률-로데이터 입력,변환_양지1센터.xlsx"
KGA2_PATH = BASE_DIR / "data/master/가동률-로데이터 입력,변환 양지2센터.xlsx"
KGA1_TEMP = BASE_DIR / "data/temp/tmp_kga1.xlsx"
KGA2_TEMP = BASE_DIR / "data/temp/tmp_kga2.xlsx"

LOC_MASTER1 = BASE_DIR / "data/master/양지1센터_로케이션_정보.xlsx"
LOC_MASTER2 = BASE_DIR / "data/master/양지2센터_로케이션_정보.xlsx"
LOC_MASTER3 = BASE_DIR / "data/master/양지3센터_로케이션_정보.xlsx"
PRICE_MASTER = BASE_DIR / "data/master/기준정보_공장도가.xlsx"

SUPABASE_URL = os.getenv("SUPABASE_POOLER_URL", os.getenv("SUPABASE_DB_URL", ""))

# ── 상수 ─────────────────────────────────────────────────────────────
_REGION_RE = re.compile(r'경\d{2}-|광주\d{2}-|전남\d{3}-')
_3PL_EXCL  = {
    # 수동 다운로드 파일 기준 (한국어 오너명)
    "퍼시스", "일룸", "데스커", "시디즈", "알로소", "슬로우베드",
    # RPA API 다운로드 기준 (WMS 오너 코드)
    "T60I01", "T60I02", "T60I03",   # 일룸/슬로우베드
    "T60F01",                        # 퍼시스
    "T60P01", "T60P02",              # 시디즈(T60P01), 알로소(T60P02)
}
_VALID_Z1  = {"A-P", "B", "C-D", "E-F", "H-I", "J-K", "L", "L/S", "P/S"}
_ZONE2_MAP = {"M": "M-N", "N": "M-N", "S": "S"}
_ZONE3_MAP = {"W": "W", "R": "R", "T1": "R", "T3": "R", "T4": "R"}

# 종합실적 D열 위치 (1-indexed): (표준시간 행, 실적시간 행)
ZONE_ROWS_1 = {
    "H-I": (10,  21),  "C-D": (34,  45),  "A-P": (58,  69),
    "DPS": (106, 117),
    "E-F": (188, 199), "J-K": (212, 223), "L":   (236, 247),
    "B":   (260, 271), "L/S": (284, 295),
}
ZONE_ROWS_2 = {
    "M-N": (10,  21), "S": (34,  45),
    "W":   (188, 199), "R": (212, 223),
}

# 피킹실적 C열 작업자 슬롯 (1-indexed, 양끝 포함)
PICKING_SLOTS_1 = {
    "H-I": (7,   56),  "C-D": (58,  107), "A-P": (109, 158),
    "DPS": (167, 216),
    "E-F": (227, 276), "J-K": (278, 327), "L":   (329, 378),
    "B":   (380, 429), "L/S": (431, 480),
}
PICKING_SLOTS_2 = {
    "M-N": (7,   56),  "S":   (58,  107),
    "W":   (118, 167), "R":   (169, 218),
}

ZONE_OWNER = {
    "H-I": "일룸",   "C-D": "일룸",   "A-P": "일룸",   "DPS": "일룸",
    "E-F": "퍼시스", "J-K": "퍼시스", "L":   "퍼시스", "B":   "퍼시스", "L/S": "퍼시스",
    "M-N": "데스커", "S":   "데스커",
    "W":   "3PL",    "R":   "3PL",
}

# ── 브랜드별 휴게시간 ─────────────────────────────────────────────────
# excel_deduct: 수식이 이미 공제하는 분. 0이면 수식 공제 없음.
# check_from: 이력 체크 시작 시간 (None이면 start와 동일). 저녁만 17:40.
# date_ref: 날짜 기준 — "start"=wave_start날짜, "end"=wave_end날짜 (자정이후시간대)
from datetime import time as _dtime
_B = lambda name, s, e, mins, chk=None, xl=0, dr="start": {
    "name": name, "start": s, "end": e, "minutes": mins,
    "check_from": chk, "excel_deduct": xl, "date_ref": dr,
}
BREAK_TIMES = {
    "일룸": {
        "주간": [
            _B("점심", _dtime(12,10), _dtime(13,10), 60, xl=50),
            _B("저녁", _dtime(17,30), _dtime(18, 0), 30, chk=_dtime(17,40), xl=30),
        ],
        "야간": [
            _B("야식", _dtime( 1, 0), _dtime( 2, 0), 60, dr="end"),
        ],
    },
    "퍼시스": {
        "주간": [
            _B("점심", _dtime(12,20), _dtime(13,10), 50, xl=50),
            _B("저녁", _dtime(17,30), _dtime(18, 0), 30, chk=_dtime(17,40), xl=30),
        ],
        "야간": [],
    },
    "데스커": {
        "주간": [
            _B("점심", _dtime(12,20), _dtime(13,20), 60, xl=50),
            _B("저녁", _dtime(17,30), _dtime(18, 0), 30, chk=_dtime(17,40), xl=30),
        ],
        "야간": [
            _B("야식", _dtime( 1, 0), _dtime( 2, 0), 60, dr="end"),
        ],
    },
    "3PL": {
        "주간": [
            _B("점심", _dtime(12,10), _dtime(13, 0), 50, xl=50),
            _B("저녁", _dtime(17,30), _dtime(18, 0), 30, chk=_dtime(17,40), xl=30),
        ],
        "야간": [],
    },
}
ZONE_CENTER = {
    "H-I": "양지1센터", "C-D": "양지1센터", "A-P": "양지1센터", "DPS": "양지1센터",
    "E-F": "양지1센터", "J-K": "양지1센터", "L":   "양지1센터", "B":   "양지1센터", "L/S": "양지1센터",
    "M-N": "양지2센터", "S":   "양지2센터",
    "W":   "양지3센터", "R":   "양지3센터",
}

# 검증 기준값 — 날짜별 (20260610_기준정보_마스터.xlsx AU열 직접 집계 기준)
EXPECTED_BY_DATE = {
    # 2026-05-12: raw 모드 재계산 (G-zone fix + NaT fix 최종 반영)
    "2026-05-12": {
        "H-I": 17.4550, "C-D": 40.7895, "A-P": 21.4249, "DPS": 48.9373,
        "E-F": 17.6566, "J-K": 19.6594, "L":   5.8695,  "B":   1.5881,  "L/S": 5.6261,
        "M-N": 31.7577, "S":   15.1104, "W":   3.2505,  "R":   3.5990,
    },
    # 2026-05-13: 재계산 필요 (참고용)
    "2026-05-13": {
        "H-I": 19.1706, "C-D": 36.4978, "A-P": 21.3820, "DPS": 51.2507,
        "E-F": 24.1999, "J-K": 16.4616, "L":   6.7264,  "B":   3.6375, "L/S": 4.9173,
        "M-N": 34.4382, "S":   14.8258, "W":   9.7850,  "R":   5.9627,
    },
    # 2026-06-08: 20260608_기준정보_마스터.xlsx 종합실적 D열 직접 읽음
    "2026-06-08": {
        "H-I": 24.0598, "C-D": 48.6436, "A-P": 30.2567, "DPS": 61.6895,
        "E-F": 18.1599, "J-K": 15.8896, "L":   3.2653,  "B":   1.9556,  "L/S": 3.4598,
        "M-N": 33.3281, "S":   16.3061, "W":   6.4689,  "R":   8.5688,
    },
    # 2026-06-09: 20260609_기준정보_마스터.xlsx 종합실적 D열 직접 읽음
    "2026-06-09": {
        "H-I": 25.4219, "C-D": 47.1305, "A-P": 26.2099, "DPS": 66.6465,
        "E-F": 13.7959, "J-K": 11.8914, "L":   4.7090,  "B":   2.5222,  "L/S": 4.0178,
        "M-N": 32.7267, "S":   15.2149, "W":   5.3717,  "R":   4.8839,
    },
    # 2026-06-10: 20260610_기준정보_마스터.xlsx AU열 직접 집계 (이 값이 절대 기준)
    "2026-06-10": {
        "H-I": 27.0850, "C-D": 42.7259, "A-P": 26.2093, "DPS": 58.2184,
        "E-F": 17.7274, "J-K": 15.7949, "L":   7.8425,  "B":   1.9487,  "L/S": 4.1044,
        "M-N": 33.4419, "S":   16.8485, "W":   5.7830,  "R":   7.2018,
    },
    # 2026-06-11: 20260611_기준정보_마스터.xlsx 종합실적 D열 직접 읽음
    #   ※ DPS: 마스터는 수식을 1738행까지만 채우고 마지막 670행(10001~10670)의
    #     M~AU 수식을 누락 → 종합실적 69.2564는 표준시간이 과소된 불완전 값.
    #     데이터(A~L)는 2408행 모두 존재하므로, 전체 수식 적용한 86.9662가 정확.
    #     (DPS는 외부 API 입력이라 수식 그대로 계산 — 사용자 확인 2026-06-25)
    "2026-06-11": {
        "H-I": 31.3740, "C-D": 56.8500, "A-P": 32.4312, "DPS": 86.9662,
        "E-F": 31.1024, "J-K": 18.2566, "L":   7.9756,  "B":   3.2166,  "L/S": 6.3628,
        "M-N": 40.2882, "S":   20.5462, "W":   6.4017,  "R":   8.2060,
    },
    # 2026-06-12: 20260612_기준정보_마스터.xlsx 종합실적 D열 직접 읽음
    "2026-06-12": {
        "H-I": 32.4042, "C-D": 51.7323, "A-P": 30.2481, "DPS": 76.1152,
        "E-F": 29.8744, "J-K": 20.7815, "L":   7.2382,  "B":   2.7559,  "L/S": 6.4319,
        "M-N": 39.3831, "S":   19.2881, "W":   3.7202,  "R":   5.1842,
    },
}

# ── 절대 기준: 종합실적_최종 파일에서 날짜별 직접 읽기 ──────────────────
# daily 마스터 종합실적보다 우선. (특히 DPS는 daily가 수식누락으로 과소 → 최종파일이 정확)
FINAL_PATH = BASE_DIR / "data/master/2026년 06월 피킹 종합실적_최종.xlsx"
# '종합실적' 시트(3번째)의 구역별 '표준시간(작업시간)[hr]' 행 (1-indexed)
ZONE_STD_ROW_FINAL = {
    "H-I": 10,  "C-D": 34,  "A-P": 58,  "DPS": 106,
    "E-F": 190, "J-K": 214, "L":   238, "B":   262, "L/S": 286,
    "M-N": 355, "S":   379, "W":   519, "R":   543,
}
_final_exp_cache: dict = {}

def _load_expected_from_final(date_str: str):
    """종합실적_최종 '종합실적' 시트에서 해당 날짜 열의 구역별 표준시간 dict 반환.
    열 매핑: 6/1=4열, 이후 1일당 1열씩 연속(주말 포함). 2026년 6월 전용.
    """
    if date_str in _final_exp_cache:
        return _final_exp_cache[date_str]
    from datetime import date as _date
    try:
        y, m, d = map(int, date_str.split("-"))
        if (y, m) != (2026, 6) or not FINAL_PATH.exists():
            _final_exp_cache[date_str] = None
            return None
        col = 4 + (_date(y, m, d) - _date(2026, 6, 1)).days
        import openpyxl
        wb = openpyxl.load_workbook(FINAL_PATH, read_only=True, data_only=True)
        ws = wb.worksheets[2]
        out = {}
        for z, r in ZONE_STD_ROW_FINAL.items():
            v = ws.cell(r, col).value
            if isinstance(v, (int, float)) and v > 0:
                # 블록: 표준=r, 금액=r-3(총피킹금액), 박스=r-1(총피킹박스수), 실적=r+11(실적시간)
                out[z] = {
                    "std":    float(v),
                    "amount": float(ws.cell(r - 3, col).value or 0),
                    "box":    float(ws.cell(r - 1, col).value or 0),
                    "act":    float(ws.cell(r + 11, col).value or 0),
                }
        wb.close()
        _final_exp_cache[date_str] = out or None
        return out or None
    except Exception:
        _final_exp_cache[date_str] = None
        return None


# ── 로케이션 마스터 로드 ──────────────────────────────────────────────
_loc_cache: dict = {}

def _load_loc_map(path: Path) -> dict:
    key = str(path)
    if key in _loc_cache:
        return _loc_cache[key]
    df = pd.read_excel(path, usecols=[2, 3], header=0)
    df.columns = ["zone", "loc_id"]
    m = {
        str(r.loc_id).strip(): str(r.zone).strip()
        for r in df.itertuples()
        if pd.notna(r.loc_id) and pd.notna(r.zone)
    }
    _loc_cache[key] = m
    return m


# ── zone 매핑 ────────────────────────────────────────────────────────
# lm1 미등록 위치(신규 랙 등)를 위한 prefix 기반 fallback
# (비교 분석 결과: 마스터에는 있고 우리에 없는 행이 lm 미등록 위치에 집중됨)
_Z1_PREFIX_FALLBACK = {
    "A": "A-P",
    "B": "B",
    "C": "C-D", "D": "C-D",
    "E": "E-F",  "F": "E-F",
    "G": "J-K",  # G랙(G-1xx 등)은 J-K zone (퍼시스 소물)
    "H": "H-I",  "I": "H-I",
    "J": "J-K",  "K": "J-K",
    "L": "L",    # L/S는 lm1에 있는 위치만 해당 → 미등록 L-XXX는 L로 처리
}

# 로케이션 마스터 오류 보정: 마스터에 잘못 등록된 zone 값 override
_Z1_ZONE_OVERRIDE: dict[str, str] = {
    "F-102-26":  "E-F",  # 마스터 오류: B로 분류됨, 실제 F구역 → E-F
    "L-103-05-4": "L",  # 마스터 오류: L/S로 분류됨, 실제 L구역
    "L-103-16-4": "L",  # 마스터 오류: L/S로 분류됨, 실제 L구역
}


def _zone1(loc: str, lm: dict) -> str | None:
    """양지1센터: 일룸(H-I,C-D,A-P,DPS) + 퍼시스(E-F,J-K,L,B,L/S)"""
    if not loc or not isinstance(loc, str):
        return None
    loc = loc.strip()
    if re.match(r'^Y-REC', loc, re.IGNORECASE):
        return None
    override = _Z1_ZONE_OVERRIDE.get(loc)
    if override:
        return override
    if re.match(r'^P-110', loc, re.IGNORECASE):
        return "A-P"  # P-110은 단가상 H-I이나 생산성 기준으로는 A-P로 처리
    m = re.match(r'^P-(\d+)', loc, re.IGNORECASE)
    if m and int(m.group(1)) >= 300:
        return "P/S"
    raw = lm.get(loc)
    if not raw:
        # lm1 미등록: prefix 기반 fallback (L-106-08-1 → "L" 등)
        prefix = loc.split("-")[0].upper() if "-" in loc else loc[0].upper()
        fb = _Z1_PREFIX_FALLBACK.get(prefix)
        if fb:
            return fb
        # P-XXX (XXX < 300, 110 이외): A-P fallback
        if re.match(r'^P-(\d+)', loc, re.IGNORECASE):
            return "A-P"
        return None
    cleaned = re.sub(r'\(반품\)', '', raw).strip()
    if cleaned in _VALID_Z1:
        return cleaned
    # lm1값이 단순 prefix 문자(예: "G")여서 유효 zone이 아닌 경우 → prefix fallback 시도
    # 예: lm1["G-100-04-1"] = "G" → _Z1_PREFIX_FALLBACK["G"] = "J-K"
    prefix = cleaned[0].upper() if cleaned else (loc.split("-")[0].upper() if "-" in loc else loc[0].upper())
    fb = _Z1_PREFIX_FALLBACK.get(prefix)
    return fb  # None이면 None 반환


def _zone2(loc: str, lm: dict) -> str | None:
    """양지2센터: 데스커(M-N, S)"""
    if not loc or not isinstance(loc, str):
        return None
    loc = loc.strip()
    raw = lm.get(loc)
    if not raw:
        # lm2 미등록: prefix 기반 fallback (M-118-XX → "M-N" 등)
        prefix = loc.split("-")[0].upper() if "-" in loc else loc[0].upper()
        if prefix in ("M", "N"):
            return "M-N"
        if prefix == "S":
            return "S"
        return None
    cleaned = re.sub(r'\(반품\)|\(보관\)', '', raw).strip()
    return _ZONE2_MAP.get(cleaned)


def _zone3(loc: str, lm: dict) -> str | None:
    """양지3센터: 3PL(W, R)"""
    if not loc or not isinstance(loc, str):
        return None
    raw = lm.get(loc.strip())
    if not raw:
        return None
    return _ZONE3_MAP.get(raw.strip())


# ── raw 파일 탐색 ────────────────────────────────────────────────────
def find_raw(target: date, owner: str) -> Path | None:
    mmdd = target.strftime("%m%d")
    nxdd = (target + timedelta(days=1)).strftime("%m%d")
    d = RAW_DIR / target.strftime("%Y/%m")
    patterns = {
        "퍼시스": [f"퍼시스_{mmdd}.xlsx"],
        "일룸":   [f"일룸_{mmdd}_{nxdd}.xlsx", f"일룸_{mmdd}*.xlsx"],
        "데스커": [f"데스커_{mmdd}_{nxdd}.xlsx", f"데스커_{mmdd}*.xlsx"],
        "3PL":    [f"3PL_{mmdd}.xlsx", f"3PL_{mmdd}*.xlsx", f"3센터_{mmdd}.xlsx"],
    }
    for pat in patterns.get(owner, []):
        hits = sorted(glob.glob(str(d / pat)))
        if hits:
            return Path(hits[0])
    return None


# ── raw 로드 / 필터 ──────────────────────────────────────────────────
def _load_raw(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path)
    # 앞뒤 공백 제거 + 내부 공백 언더스코어로 통일 ('ITEM ID' → 'ITEM_ID', 'PLT ID' → 'PLT_ID')
    df.columns = df.columns.str.strip().str.replace(r'\s+', '_', regex=True)
    # 작업일시 파싱 문제 수정:
    # '2026-05-12T11:21:05.907' (밀리초 O) vs '2026-05-12T11:21:05' (밀리초 X) 혼재 시
    # pandas가 밀리초 포함 형식을 열 전체에 고정 추론하여 밀리초 없는 행을 NaT 처리.
    # → format="mixed" (pandas 2.0+): 각 값을 개별 파싱
    # → 구버전 fallback: 1차 파싱 후 NaT 행만 재파싱
    raw_col = df["작업일시"].astype(str).str.replace("T", " ", regex=False)
    try:
        df["작업일시"] = pd.to_datetime(raw_col, errors="coerce", format="mixed")
    except TypeError:
        # pandas < 2.0: format="mixed" 미지원
        df["작업일시"] = pd.to_datetime(raw_col, errors="coerce")
        nat_mask = df["작업일시"].isna()
        if nat_mask.any():
            # NaT 행만 재시도 (infer_datetime_format=False로 형식 고정 방지)
            df.loc[nat_mask, "작업일시"] = pd.to_datetime(
                raw_col[nat_mask], errors="coerce", infer_datetime_format=False
            )
    return df.dropna(subset=["작업일시"])


def _parse_file_end_date(path, t: date) -> date:
    """파일명 마지막 _NNDD 토큰에서 종료 날짜 파싱.

    예) 일룸_0602_0604.xlsx → 2026-06-04
        일룸_0601_0602.xlsx → 2026-06-02 (평상시)
        일룸_0605_0608.xlsx → 2026-06-08 (연속공휴일 후)
    파싱 실패 시 t+1(평상시 기본값) 반환.
    """
    if path is None:
        return t + timedelta(days=1)
    stem = Path(path).stem  # e.g., '일룸_0602_0604'
    parts = stem.split('_')
    if len(parts) >= 3:
        end_mmdd = parts[-1]
        if len(end_mmdd) == 4:
            try:
                m, d = int(end_mmdd[:2]), int(end_mmdd[2:])
                return date(t.year, m, d)
            except (ValueError, TypeError):
                pass
    return t + timedelta(days=1)


def _i1_d1_window(t: date, raw_path):
    """일룸/데스커 주간·야간 타임스탬프 범위 반환 (day_start, day_end, night_start, night_end).

    파일명의 종료 날짜(NNDD)로 야간 윈도우 결정:
    - 평상시 NNDD = t+1  → 야간 = t 21:00 ~ (t+1) 08:00
    - 공휴일 전날 NNDD > t+1 → 야간 = (NNDD-1) 21:00 ~ NNDD 08:00
      예) 일룸_0602_0604.xlsx: 야간 = 6/3 21:00 ~ 6/4 08:00 (6/3 공휴일 야간투입)
          일룸_0605_0608.xlsx: 야간 = 6/7 21:00 ~ 6/8 08:00 (연속공휴일 후 야간투입)
    """
    ts = pd.Timestamp(t)
    day_start   = ts.replace(hour=8,  minute=0,  second=0)
    day_end     = ts.replace(hour=20, minute=59, second=59)
    end_date    = _parse_file_end_date(raw_path, t)
    ts_end      = pd.Timestamp(end_date)
    ts_prev     = ts_end - pd.Timedelta(days=1)
    night_start = ts_prev.replace(hour=21, minute=0, second=0)
    night_end   = ts_end.replace(hour=8,  minute=0, second=0)
    return day_start, day_end, night_start, night_end


def _filter_f1(df: pd.DataFrame, t: date, raw_path=None) -> pd.DataFrame:
    """퍼시스: 주간(08:01~20:59), Y-REC 제외"""
    ts = pd.Timestamp(t)
    df = df[df["작업일시"].between(
        ts.replace(hour=8, minute=1, second=0),
        ts.replace(hour=20, minute=59, second=59),
    )]
    return df[~df["LOCATION"].astype(str).str.upper().str.startswith("Y-REC")]


def _filter_i1(df: pd.DataFrame, t: date, raw_path=None) -> pd.DataFrame:
    """일룸: 주간 + 야간, Y-REC 제외, [주간]/[야간] 태그 필수 (DPS P-3XX 예외).

    파일명 종료 날짜(_NNDD)로 야간 윈도우 동적 결정 — _i1_d1_window 참조.
    """
    day_start, day_end, night_start, night_end = _i1_d1_window(t, raw_path)
    mask = df["작업일시"].between(day_start, day_end) | df["작업일시"].between(night_start, night_end)
    out  = df[mask & ~df["LOCATION"].astype(str).str.upper().str.startswith("Y-REC")].copy()
    tagged  = out["작업자"].astype(str).str.match(r"^\[(주간|야간)\]")
    dps_loc = out["LOCATION"].astype(str).apply(
        lambda loc: bool(re.match(r"^P-(\d+)", loc, re.I))
                    and int(re.match(r"^P-(\d+)", loc, re.I).group(1)) >= 300
    )
    # DPS(P-3XX)는 주간만 포함 — 야간 DPS는 다음날 업무일자에 귀속
    dps_day = out["작업일시"].between(day_start, day_end)
    return out[tagged | (dps_loc & dps_day)]


def _filter_d1(df: pd.DataFrame, t: date, raw_path=None) -> pd.DataFrame:
    """데스커: 주간 + 야간, Y-REC 제외.

    파일명 종료 날짜(_NNDD)로 야간 윈도우 동적 결정 — _i1_d1_window 참조.
    야간 조:
    - [야간] 태그: night_start - 1h(20:00)부터 허용 (얼리스타트)
    - [주간] 등 태그 없음: night_start(21:00) 이후라면 야간 연장으로 인정
      (WMS에서 [야간] 태그 없이 21시 이후 작업이 기록되는 경우 포함)
    """
    day_start, day_end, night_start, night_end = _i1_d1_window(t, raw_path)
    night_start_early = night_start - pd.Timedelta(hours=1)
    no_yrec = ~df["LOCATION"].astype(str).str.upper().str.startswith("Y-REC")
    is_night_tag = df["작업자"].astype(str).str.match(r"^\[야간\]")
    # 주간: 시간대 내에서 [야간] 태그가 없으면 포함
    mask_day = df["작업일시"].between(day_start, day_end) & ~is_night_tag
    # 야간: [야간] 태그는 얼리스타트(20:00)부터, 무태그는 21:00부터 포함
    mask_night = (
        (df["작업일시"].between(night_start_early, night_end) & is_night_tag) |
        (df["작업일시"].between(night_start, night_end) & ~is_night_tag)
    )
    return df[(mask_day | mask_night) & no_yrec].copy()


def _filter_du1(df: pd.DataFrame, t: date, raw_path=None) -> pd.DataFrame:
    """3PL: 주간(08:01~20:59), 그룹사(퍼시스/일룸 등) OWNER 제외, Y-REC 제외.

    날짜 경계를 두지 않고 시간대만 필터 — RPA가 다운로드 범위를
    (target ~ 다음근무일-1)로 이미 제한하므로 파일 내 모든 날짜가
    주말·공휴일 합산 대상이 될 수 있음.
    그룹사리스트(_3PL_EXCL)는 KGA 파일에 넣을 때 제거한다 — 기준 마스터 기준 R=7.2018hr 일치 확인.
    """
    h = df["작업일시"].dt.hour
    m = df["작업일시"].dt.minute
    in_window = ((h > 8) | ((h == 8) & (m >= 1))) & (h <= 20)
    df = df[in_window]
    df = df[~df["LOCATION"].astype(str).str.upper().str.startswith("Y-REC")]
    return df[~df["OWNER"].astype(str).isin(_3PL_EXCL)]


# ── 위치 보정 테이블 (raw 데이터 입력 오류 보정) ────────────────────
# I_1 (일룸) 전용: B구역 위치는 C-D zone 치환
# "일룸에 B구역이 저거 하나" — B-109-01-1 피킹은 실제 C-103-01-1 위치
_I1_LOC_CORRECTIONS: dict[str, str] = {
    "B-109-01-1": "C-103-01-1",
}

# ── 시트 데이터 빌드 ─────────────────────────────────────────────────
def _to_int_str(v) -> str:
    """숫자형 .0 suffix 제거 ('55.0'→'55', '202606010227.0'→'202606010227').
    문자열은 그대로 반환 ('00'→'00', leading zero 보존).
    """
    if v is None:
        return ''
    if isinstance(v, float):
        if pd.isna(v):
            return ''
        if v == int(v):
            return str(int(v))
        return str(v)
    if isinstance(v, int):
        return str(v)
    return str(v)  # str 그대로 ('00' → '00')


def _build(df: pd.DataFrame, zone_fn, fixed_region: str | None = None,
           loc_corrections: dict | None = None,
           last_zones: list | None = None) -> pd.DataFrame:
    """
    raw 행 → 시트 레코드 변환 + 정렬

    정렬 기준 (가동률-로데이터 입력,변환 파일 기준):
      일반 구역: J(작업자) > K(작업일시) > I(WAVE명) > G(PLT_ID) > F(LOCATION)
      DPS:       J(작업자) > I(WAVE명)   > G(PLT_ID) > K(작업일시) > F(LOCATION)

    last_zones: 마스터에서 항상 맨 뒤에 붙는 구역 목록
      F_1=['L/S'], I_1=['DPS'], D_1=['S'], DU_1=[]
    """
    if last_zones is None:
        last_zones = ['DPS']

    records = []
    for _, row in df.iterrows():
        loc  = str(row.get("LOCATION", ""))
        if loc_corrections and loc in loc_corrections:
            loc = loc_corrections[loc]
        zone = zone_fn(loc)
        if zone is None:
            continue
        worker    = str(row.get("작업자", ""))
        wave_name = str(row.get("WAVE명", ""))
        region    = fixed_region or ("지방" if _REGION_RE.search(wave_name) else "소액")
        if zone == "P/S":
            zone   = "DPS"
            worker = "DPS"
        ts = row["작업일시"]
        hour   = ts.hour   if isinstance(ts, (pd.Timestamp, datetime)) else 0
        minute = ts.minute if isinstance(ts, (pd.Timestamp, datetime)) else 0
        second = ts.second if isinstance(ts, (pd.Timestamp, datetime)) else 0
        records.append({
            "A": zone + worker,
            "B": zone,
            "C": row.get("오더번호", ""),
            "D": row.get("ITEM ID", row.get("ITEM_ID", "")),
            "E": row.get("피킹수량", 0),
            "F": loc,
            "G": _to_int_str(row.get("PLT ID", row.get("PLT_ID", ""))),
            "H": _to_int_str(row.get("WAVE번호", "")),
            "I": wave_name,
            "J": worker,
            "K": ts,
            "L": region,
            "Y": hour,
            "Z": minute,
            "AA": second,
        })
    if not records:
        return pd.DataFrame(columns=list("ABCDEFGHIJKL") + ["Y", "Z", "AA"])

    out  = pd.DataFrame(records)
    last = out[out["B"].isin(last_zones)]
    rest = out[~out["B"].isin(last_zones)]

    # 정렬 기준 (docstring 참조):
    #   일반: J(작업자) > K(작업일시) > I(WAVE명) > G(PLT_ID) > F(LOCATION)
    #   DPS : J(작업자) > I(WAVE명)   > G(PLT_ID) > K(작업일시) > F(LOCATION)
    rest_sorted = rest.sort_values(["J", "K", "I", "G", "F"])
    # DPS(last): J="DPS"로 통일 후 I>G>K>F 정렬 (AK 이동시간이 I=WAVE명 경계로 발화)
    last_dps  = last[last["B"] == "DPS"].copy()
    last_dps["J"] = "DPS"
    last_rest = last[last["B"] != "DPS"]
    last_sorted = pd.concat([
        last_rest.sort_values(["J", "K", "I", "G", "F"]),
        last_dps.sort_values(["J", "I", "G", "K", "F"]),
    ])

    return pd.concat([rest_sorted, last_sorted]).reset_index(drop=True)


# ── pywin32 유틸 ─────────────────────────────────────────────────────
_XL_ERRORS = frozenset({
    -2146826246, -2146826252, -2146826259, -2146826265,
    -2146826273, -2146826281, -2146826288,
})


def _safe(v) -> float:
    if not isinstance(v, (int, float)):
        return 0.0
    if isinstance(v, float) and v != v:
        return 0.0
    if isinstance(v, int) and v in _XL_ERRORS:
        return 0.0
    return float(v)


def _df_to_rows(df: pd.DataFrame) -> list:
    out = []
    for row in df[list("ABCDEFGHIJKL")].itertuples(index=False):
        clean = []
        for v in row:
            try:
                if pd.isna(v):
                    clean.append(None)
                    continue
            except (TypeError, ValueError):
                pass
            if isinstance(v, pd.Timestamp):
                clean.append(v.to_pydatetime().replace(tzinfo=None))
            else:
                clean.append(v)
        out.append(clean)
    return out


_WRITE_CHUNK = 5000  # COM Value= 대량 배열 한계(~9999행) 우회


def _write_sheet(ws, df: pd.DataFrame):
    """A~L열 데이터 초기화 후 청크 쓰기.

    win32com Range.Value= 단일 배열이 ~9999행에서 잘리는 현상을 방지하기 위해
    _WRITE_CHUNK 행 단위로 분할 작성.
    """
    last = ws.UsedRange.Row + ws.UsedRange.Rows.Count - 1
    if last >= 2:
        ws.Range(f"A2:L{last}").ClearContents()
    rows = _df_to_rows(df)
    for i in range(0, len(rows), _WRITE_CHUNK):
        chunk = rows[i:i + _WRITE_CHUNK]
        sr = 2 + i
        er = sr + len(chunk) - 1
        ws.Range(ws.Cells(sr, 1), ws.Cells(er, 12)).Value = \
            tuple(tuple(r) for r in chunk)
    print(f"    [{ws.Name}] {len(df)}행 작성")


def _write_v_col(ws, df: pd.DataFrame, t3t4_cap: bool = False):
    """
    V(22)/W(23)/X(24)열: LOCATION(F열)에서 랙번호·베이번호·단높이를 추출하여 쓰기.

    마스터의 W/X 열은 수식이 아닌 정적 값으로 저장되어 있어,
    이전 실행의 데이터 값이 그대로 잔류한다.
    특히 DU_1 T-3xx 위치(T3/T4 zone)는 X=4, W=9 같은 stale 값이 고착되어
    AM(단높이 준비시간) 계산이 크게 왜곡된다.
    → V와 함께 W, X도 매번 올바른 값으로 덮어씌워야 한다.

    t3t4_cap=True (DU_1 전용):
      T3/T4 zone 위치(rack ≥ 300)는 거리 테이블에 해당 번호가 없으므로 V=101 고정.
      X=4 고정: AM 계산(lvl_10_2[4]=1.66min)이 마스터와 일치해야 함.
      W는 실제 값 그대로 사용.
    """
    n = len(df)
    if n == 0:
        return
    vwx_data = []
    for loc in df["F"]:
        parts = str(loc).split("-")
        try:
            rack = int(parts[1])
            if t3t4_cap and rack >= 300:
                rack = 101
        except (IndexError, ValueError):
            rack = None
        try:
            bay = int(parts[2])
        except (IndexError, ValueError):
            bay = 1
        try:
            level_digit = int(parts[3])
            if t3t4_cap and rack == 101:  # T3/T4 zone (DU_1 rack≥300) → X=4 고정
                raw_level = 4
            else:
                zone_ch = str(loc)[0] if loc else ''
                sheet_name = ws.Name  # 'F_1', 'I_1', 'D_1', 'DU_1'
                if sheet_name == 'I_1':
                    # 일룸: H/B zone → 마지막자리, 기타 → bay>25이면 4
                    raw_level = level_digit if zone_ch in ('H', 'B') else (4 if bay > 25 else level_digit)
                elif sheet_name == 'F_1':
                    # 퍼시스: bay>24이면 4 (기준정보 1~24까지만)
                    raw_level = 4 if bay > 24 else level_digit
                else:
                    # D_1(데스커) / DU_1(3PL): rack>200→4 (T3T4는 위에서 처리)
                    raw_level = 4 if (rack is not None and rack > 200) else level_digit
            # 마스터 1_10 시트: 1,2=정수(0분), 3+=L접두사(L3=0,L4=1.36,L5=1.36...)
            level = raw_level if raw_level <= 2 else f'L{raw_level}'
        except (IndexError, ValueError):
            raw_level = 4 if (t3t4_cap and rack == 101) else 1
            level = raw_level if raw_level <= 2 else f'L{raw_level}'
        vwx_data.append([rack, bay, level])
    # V(22), W(23), X(24) 동시 갱신
    ws.Range(ws.Cells(2, 22), ws.Cells(n + 1, 24)).Value = \
        tuple(tuple(r) for r in vwx_data)
    print(f"    [{ws.Name}] V/W/X열(랙/베이/단) {n}행 갱신")


def _write_u_col(ws, df: pd.DataFrame):
    """U(21)열: LOCATION(F열) 첫 문자(랙 타입)를 쓰기.

    마스터 U열은 수식이 아닌 정적 값으로 이전 실행 데이터가 잔류한다.
    현재 데이터 행 수가 마스터 잔류 범위보다 많으면 나머지 행의 U=None이 되어
    AD(30)/AG(33) MATCH($U, ABCD!$B$2) 조회 실패 → AS/AT/AU #N/A 전파 발생.
    → 매 실행마다 올바른 값으로 덮어써야 한다.

    zone "L/S" → "L/S", zone "DPS" → "P", 나머지 → location 첫 문자 대문자
    """
    n = len(df)
    if n == 0:
        return
    u_vals = []
    for _, row in df.iterrows():
        zone = str(row.get("B", ""))
        if zone == "L/S":
            u = "L/S"
        elif zone == "DPS":
            u = "P"
        else:
            parts = str(row.get("F", "")).split("-")
            u = parts[0].upper() if parts and parts[0] else ""
        u_vals.append((u,))
    ws.Range(ws.Cells(2, 21), ws.Cells(n + 1, 21)).Value = tuple(u_vals)
    print(f"    [{ws.Name}] U열(랙 타입) {n}행 갱신")


def _write_yzaa_cols(ws, df: pd.DataFrame):
    """Y(25)/Z(26)/AA(27)열: 작업일시(K열)에서 시/분/초 추출하여 쓰기."""
    n = len(df)
    if n == 0:
        return
    yzaa_data = []
    for _, row in df.iterrows():
        ts = row.get("K")
        if isinstance(ts, (pd.Timestamp, datetime)):
            yzaa_data.append((ts.hour, ts.minute, ts.second))
        else:
            yzaa_data.append((0, 0, 0))
    ws.Range(ws.Cells(2, 25), ws.Cells(n + 1, 27)).Value = tuple(yzaa_data)
    print(f"    [{ws.Name}] Y/Z/AA열(시/분/초) {n}행 갱신")


# ── 가동률-로데이터 파일 COM 파이프라인 ──────────────────────────────

_XL_EPOCH = datetime(1899, 12, 30)

def _to_excel_serial(dt):
    """python datetime/Timestamp → Excel serial 날짜 숫자.
    win32com에 datetime 객체를 직접 .Value로 쓰면 KST(-9h) 시프트가 발생하여
    08:40 → 전날 23:40로 깨짐(검증됨). serial 숫자로 쓰면 시프트 없음."""
    if dt is None:
        return None
    if isinstance(dt, pd.Timestamp):
        dt = dt.to_pydatetime()
    if not isinstance(dt, datetime):
        return dt
    dt = dt.replace(tzinfo=None)
    delta = dt - _XL_EPOCH
    return delta.days + (delta.seconds + delta.microseconds / 1e6) / 86400.0


def _fill_raw_data_sheet(ws_raw, raw_df: pd.DataFrame):
    """가동률-로데이터 파일 RAW DATA 시트에 필터링된 raw_df 데이터 씀.
    1행(헤더) 유지, 2행부터 기존 데이터 삭제 후 새 데이터 씀.

    RAW DATA 컬럼 순서(1~12):
      Order No2, 건명, Item Id, 수량, From 로케이션, To 로케이션,
      From 팔레트, To 팔레트, Wave No, Wave 명, 작업자, 작업일시
    """
    last = ws_raw.UsedRange.Row + ws_raw.UsedRange.Rows.Count - 1
    if last >= 2:
        ws_raw.Range(ws_raw.Cells(2, 1), ws_raw.Cells(last, 12)).ClearContents()

    n = len(raw_df)
    if n == 0:
        print(f"    [{ws_raw.Name}] 데이터 없음")
        return

    rows = []
    for _, row in raw_df.iterrows():
        ts = row.get("작업일시")
        # ★ win32com datetime 직접쓰기는 -9h(KST) 시프트 발생 → Excel serial 숫자로 기입
        ts_val = _to_excel_serial(ts) if isinstance(ts, (pd.Timestamp, datetime)) else ts
        rows.append([
            row.get("오더번호", ""),
            "",                                       # 건명 (없음)
            row.get("ITEM_ID", ""),
            row.get("피킹수량", 0),
            row.get("LOCATION", ""),
            "",                                       # To 로케이션 (없음)
            "",                                       # From 팔레트 (없음)
            _to_int_str(row.get("PLT_ID", "")),
            _to_int_str(row.get("WAVE번호", "")),
            row.get("WAVE명", ""),
            row.get("작업자", ""),
            ts_val,
        ])

    for i in range(0, n, _WRITE_CHUNK):
        chunk = rows[i:i + _WRITE_CHUNK]
        sr, er = 2 + i, 1 + i + len(chunk)
        ws_raw.Range(ws_raw.Cells(sr, 1), ws_raw.Cells(er, 12)).Value = \
            tuple(tuple(r) for r in chunk)
    print(f"    [{ws_raw.Name}] {n}행 입력")


def _read_kga_calc_to_df(ws_calc, zone_fn,
                          fixed_region: str | None = None,
                          last_zones: list | None = None,
                          loc_corrections: dict | None = None) -> pd.DataFrame:
    """데이터정리,계산 시트 C(3)~T(20)열을 읽어 마스터 입력용 DataFrame 반환.

    헤더 2행 제외, 3행부터 데이터.
    반환 컬럼: A,B,C,D,E,F,G,H,I,J,K,L,U,V,W,X,Y,Z,AA

    컬럼 매핑 (0-based index → 마스터 열):
      idx 0  (C=3):  오더번호 → C
      idx 1  (D=4):  Item Id   → D
      idx 2  (E=5):  수량       → E
      idx 3  (F=6):  LOCATION  → F
      idx 4  (G=7):  To 팔레트 → G
      idx 5  (H=8):  Wave No   → H
      idx 6  (I=9):  Wave 명   → I
      idx 7  (J=10): 작업자    → J
      idx 8  (K=11): 작업일시  → K
      idx 9  (L=12): 출고지역  → L
      idx 10 (M=13): 업무일자  → skip (마스터 M열은 수식)
      idx 11 (N=14): 랙알파벳  → U(21)
      idx 12 (O=15): 랙번호    → V(22)
      idx 13 (P=16): 베이      → W(23)
      idx 14 (Q=17): 레벨      → X(24)
      idx 15 (R=18): 시        → Y(25)
      idx 16 (S=19): 분        → Z(26)
      idx 17 (T=20): 초        → AA(27)
    """
    _EMPTY = pd.DataFrame(
        columns=list("ABCDEFGHIJKL") + ["U", "V", "W", "X", "Y", "Z", "AA"]
    )
    if last_zones is None:
        last_zones = []

    xlUp = -4162
    n_last = ws_calc.Cells(ws_calc.Rows.Count, 3).End(xlUp).Row
    n_rows = n_last - 2  # 헤더 2행 제외
    if n_rows <= 0:
        return _EMPTY

    data = ws_calc.Range(
        ws_calc.Cells(3, 3), ws_calc.Cells(n_last, 20)
    ).Value
    if not data:
        return _EMPTY

    records = []
    for row_vals in data:
        if len(row_vals) < 18:
            row_vals = list(row_vals) + [None] * (18 - len(row_vals))

        loc = str(row_vals[3]).strip() if row_vals[3] is not None else ""
        if loc_corrections and loc in loc_corrections:
            loc = loc_corrections[loc]

        zone = zone_fn(loc)
        if zone is None:
            continue

        worker = str(row_vals[7]) if row_vals[7] is not None else ""
        wave_name = str(row_vals[6]) if row_vals[6] is not None else ""
        region = row_vals[9]
        if region is None or str(region).strip() == "":
            region = fixed_region or ("지방" if _REGION_RE.search(wave_name) else "소액")

        if zone == "P/S":
            zone = "DPS"
            worker = "DPS"

        ts = row_vals[8]
        if isinstance(ts, datetime) and not isinstance(ts, pd.Timestamp):
            ts = pd.Timestamp(ts)

        records.append({
            "A":  zone + worker,
            "B":  zone,
            "C":  row_vals[0],
            "D":  row_vals[1],
            "E":  row_vals[2],
            "F":  loc,
            "G":  row_vals[4],
            "H":  row_vals[5],
            "I":  wave_name,
            "J":  worker,
            "K":  ts,
            "L":  region,
            "U":  row_vals[11],
            "V":  row_vals[12],
            "W":  row_vals[13],
            "X":  row_vals[14],
            "Y":  row_vals[15],
            "Z":  row_vals[16],
            "AA": row_vals[17],
        })

    if not records:
        return _EMPTY

    df = pd.DataFrame(records)

    # K열(datetime64)을 Python datetime object로 변환 — pandas가 자동으로 datetime64로
    # 변환하는데, Python 3.13에서 itertuples/iterate 시 tzconversion 오류 발생.
    # numpy int64 (나노초) 경유로 pandas datetime 변환 코드 자체를 우회.
    import numpy as np
    from datetime import timedelta as _td
    if pd.api.types.is_datetime64_any_dtype(df["K"]):
        _NAT = np.iinfo(np.int64).min
        _EPOCH = datetime(1970, 1, 1)
        ns_vals = df["K"].values.view(np.int64)
        k_py = [
            _EPOCH + _td(microseconds=int(ns) // 1000) if ns != _NAT else None
            for ns in ns_vals
        ]
        df = df.copy()
        df["K"] = k_py
        df["K"] = df["K"].astype(object)

    last  = df[df["B"].isin(last_zones)]
    rest  = df[~df["B"].isin(last_zones)]
    rest_sorted = rest.sort_values(["J", "K", "I", "G", "F"])
    last_dps  = last[last["B"] == "DPS"].copy()
    last_dps["J"] = "DPS"
    last_rest = last[last["B"] != "DPS"]
    last_sorted = pd.concat([
        last_rest.sort_values(["J", "K", "I", "G", "F"]),
        last_dps.sort_values(["J", "I", "G", "K", "F"]),
    ])
    return pd.concat([rest_sorted, last_sorted]).reset_index(drop=True)


def _write_sheet_full(ws, df: pd.DataFrame):
    """마스터 시트에 C~L(3~12)과 U~AA(21~27) 씀.

    A(1)/B(2)는 마스터 수식이 자동 계산하므로 건드리지 않음.
    U~AA는 데이터정리,계산에서 가져온 랙/베이/레벨/시/분/초 값.
    """
    last = ws.UsedRange.Row + ws.UsedRange.Rows.Count - 1
    if last >= 2:
        ws.Range(ws.Cells(2, 1), ws.Cells(last, 12)).ClearContents()
        ws.Range(ws.Cells(2, 21), ws.Cells(last, 27)).ClearContents()

    n = len(df)
    if n == 0:
        print(f"    [{ws.Name}] 데이터 없음")
        return

    def _clean(v):
        try:
            if pd.isna(v):
                return None
        except (TypeError, ValueError):
            pass
        if isinstance(v, pd.Timestamp):
            return v.to_pydatetime().replace(tzinfo=None)
        if isinstance(v, datetime):
            dt = v.replace(tzinfo=None)
            return dt
        return v

    def _prep_cols(cols):
        """datetime64 컬럼이 itertuples에서 오류를 일으키므로
        K열(작업일시)을 Python datetime으로 미리 변환한 뒤 object dtype으로 고정."""
        sub = df[cols].copy()
        if "K" in cols:
            sub["K"] = [
                v.to_pydatetime().replace(tzinfo=None) if isinstance(v, pd.Timestamp)
                else (None if v is None else v)
                for v in sub["K"]
            ]
            sub = sub.astype({"K": object})
        return [tuple(_clean(v) for v in row) for row in sub.itertuples(index=False)]

    # A~B (col 1~2): zone·zone+작업자 직접 씀 (마스터 수식을 클리어했으므로)
    ab_rows = [(_clean(row[0]), _clean(row[1]))
               for row in df[["A", "B"]].itertuples(index=False)]
    for i in range(0, n, _WRITE_CHUNK):
        chunk = ab_rows[i:i + _WRITE_CHUNK]
        sr, er = 2 + i, 1 + i + len(chunk)
        ws.Range(ws.Cells(sr, 1), ws.Cells(er, 2)).Value = tuple(chunk)

    # C~L (col 3~12)
    cl_rows = _prep_cols(list("CDEFGHIJKL"))
    for i in range(0, n, _WRITE_CHUNK):
        chunk = cl_rows[i:i + _WRITE_CHUNK]
        sr, er = 2 + i, 1 + i + len(chunk)
        ws.Range(ws.Cells(sr, 3), ws.Cells(er, 12)).Value = tuple(chunk)

    # U~AA (col 21~27)
    uaa_rows = _prep_cols(["U", "V", "W", "X", "Y", "Z", "AA"])
    for i in range(0, n, _WRITE_CHUNK):
        chunk = uaa_rows[i:i + _WRITE_CHUNK]
        sr, er = 2 + i, 1 + i + len(chunk)
        ws.Range(ws.Cells(sr, 21), ws.Cells(er, 27)).Value = tuple(chunk)

    print(f"    [{ws.Name}] {n}행 작성 (C~L, U~AA)")


def _write_picking_workers(wb, slots: dict, sources: dict):
    """
    피킹실적 C열(작업자명) + F열(zone+작업자명 = SUMIF 키) 동시 작성.

    마스터 피킹실적 SUMIF 수식이 피킹실적!$F열을 기준키로 사용함:
      =SUMIF(D_1!$A:$A, 피킹실적!$F행, D_1!$AU:$AU)/60

    F열 원본 수식이 일부 작업자에 대해 #N/A를 반환하면 SUMIF 전체가 오류 →
    합계 셀도 오류 → 종합실적 D열 = 0. 이를 방지하기 위해 F열에 직접 값 기입.

    D_1/I_1/DU_1 A열 키 형식: zone + 작업자명 (예: "M-N[야간]김시훈")
    DPS는 "DPS" + "DPS" = "DPSDPS" (마스터 A열 동일 형식).
    """
    ws = wb.Worksheets("피킹실적")
    try:
        ws.Unprotect()
    except Exception:
        pass
    for zone, (s, e) in slots.items():
        if zone == "DPS":
            workers = ["DPS"]
        else:
            df = sources.get(zone)
            workers = list(dict.fromkeys(df["J"].dropna().tolist())) if df is not None and not df.empty else []
        size = e - s + 1
        # C열: 작업자명 (사람이 읽을 수 있는 이름)
        c_data = tuple((workers[i],) if i < len(workers) else (None,) for i in range(size))
        ws.Range(ws.Cells(s, 3), ws.Cells(e, 3)).Value = c_data
        # F열: zone+작업자명 = 해당 시트 A열과 동일한 SUMIF 키
        f_data = tuple(
            (zone + workers[i],) if i < len(workers) else (None,)
            for i in range(size)
        )
        ws.Range(ws.Cells(s, 6), ws.Cells(e, 6)).Value = f_data
        print(f"      [{zone}] {len(workers)}명 (슬롯 {size}행)")


def _read_results(ws, zone_rows: dict) -> dict:
    # 종합실적 블록: 표준=sr, 실적=ar, 박스수=sr-1(총 피킹 박스수), 금액=sr-3(총 피킹금액)
    return {
        zone: {
            "std_time_hr": _safe(ws.Cells(sr, 4).Value),
            "act_time_hr": _safe(ws.Cells(ar, 4).Value),
            "pick_count":  _safe(ws.Cells(sr - 1, 4).Value),
            "pick_amount": _safe(ws.Cells(sr - 3, 4).Value),
        }
        for zone, (sr, ar) in zone_rows.items()
    }


def _read_picking_workers(wb, slots: dict, date_str: str) -> list:
    """피킹실적 시트에서 zone별 작업자 슬롯을 읽어 작업자별 4지표 dict 리스트 반환.
       열: C(3)=작업자명, H(8)=박스수, I(9)=총피킹금액, L(12)=표준시간hr, M(13)=실적시간hr.
       (모두 우리 계산값 = SUMIF 결과. 6/1 전구역 0% 검증완료.)"""
    ws = wb.Worksheets("피킹실적")
    out = []
    for zone, (s, e) in slots.items():
        owner  = ZONE_OWNER.get(zone, "")
        center = ZONE_CENTER.get(zone, "")
        c  = ws.Range(ws.Cells(s, 3),  ws.Cells(e, 3)).Value  or []
        h  = ws.Range(ws.Cells(s, 8),  ws.Cells(e, 8)).Value  or []
        ii = ws.Range(ws.Cells(s, 9),  ws.Cells(e, 9)).Value  or []
        l  = ws.Range(ws.Cells(s, 12), ws.Cells(e, 12)).Value or []
        m  = ws.Range(ws.Cells(s, 13), ws.Cells(e, 13)).Value or []
        for i in range(e - s + 1):
            name = c[i][0] if i < len(c) and c[i] else None
            if name is None or str(name).strip() == "":
                continue
            name = str(name).strip()
            std = _safe(l[i][0]);  act = _safe(m[i][0])
            box = _safe(h[i][0]);  amt = _safe(ii[i][0])
            if std == 0 and act == 0 and box == 0 and amt == 0:
                continue
            shift = "야간" if "[야간]" in name else "주간"
            out.append({
                "work_date": date_str, "center": center, "owner": owner,
                "zone": zone, "worker_name": name, "shift": shift,
                "std_time_hr": round(std, 6), "act_time_hr": round(act, 6),
                "pick_amount": round(amt, 0) if amt else None,
                "pick_box":    int(round(box)) if box else None,
            })
    return out


def _calc_break_adjustment(wave_start_k, wave_end_k, k_times: list, brand: str, is_night: bool) -> float:
    """
    wave그룹의 휴게시간 공제 조정값 계산 (분 단위).
    반환: 양수 = 추가 차감(실적시간 감소), 음수 = 공제 취소(실적시간 증가)
    act_min에서 이 값을 빼서 적용: ax_val -= adjustment

    logic:
    - wave가 해당 휴게 시간대를 포함하면(wave_start < 휴게시작 < wave_end) 공제 검토
    - 체크 시간대(저녁: 17:40~18:00, 기타: 휴게 시작~종료)에 피킹 이력이 있으면 → 쉬지 않음
      → 수식이 이미 공제했다면 되돌림(excel_deduct만큼 +)
    - 이력이 없으면 → 쉬었음 → 수식 미공제분 추가 차감(minutes - excel_deduct)
    """
    from datetime import datetime as _dt

    if not brand or wave_start_k is None or wave_end_k is None:
        return 0.0

    # Excel win32com이 timezone-aware datetime을 반환하는 경우 naive로 변환
    def _naive(dt):
        if hasattr(dt, 'tzinfo') and dt.tzinfo is not None:
            return dt.replace(tzinfo=None)
        return dt

    wave_start_k = _naive(wave_start_k)
    wave_end_k   = _naive(wave_end_k)
    k_dts = [_naive(k) for k in k_times if k is not None and hasattr(k, 'hour')]
    if not k_dts:
        return 0.0

    breaks    = BREAK_TIMES.get(brand, {})
    shift_key = "야간" if is_night else "주간"
    break_list = breaks.get(shift_key, [])
    if not break_list:
        return 0.0

    start_date = wave_start_k.date() if hasattr(wave_start_k, 'date') else None
    end_date   = wave_end_k.date()   if hasattr(wave_end_k,   'date') else None
    if start_date is None or end_date is None:
        return 0.0

    total_adj = 0.0

    for br in break_list:
        ref_date   = start_date if br["date_ref"] == "start" else end_date
        br_start   = _dt.combine(ref_date, br["start"])
        br_end     = _dt.combine(ref_date, br["end"])
        check_from = br["check_from"]
        chk_start  = _dt.combine(ref_date, check_from) if check_from else br_start

        # wave그룹이 이 시간대를 포함하는가: wave_start < 휴게시작 < wave_end
        if not (wave_start_k < br_start < wave_end_k):
            continue

        has_work = any(chk_start <= k < br_end for k in k_dts)

        if has_work:
            # 연속 작업 → 공제 취소: 수식 공제분 되돌림
            total_adj -= br["excel_deduct"]
        else:
            # 휴게 취함 → 수식 미공제분만 추가 차감
            total_adj += br["minutes"] - br["excel_deduct"]

    return total_adj


def _python_zone_aggregate(ws, n_rows: int) -> tuple:
    """
    시트 B열(zone)·AU열(표준시간_min)·AX열(실적시간_min) 읽어 zone별 합산.
    wave그룹 단위로 휴게시간 공제를 자동 보정.

    Returns: (std_min_by_zone, act_min_by_zone) — 분(min) 단위 dict
    """
    if n_rows == 0:
        return {}, {}
    b_rng  = ws.Range(ws.Cells(2,  2), ws.Cells(n_rows + 1,  2)).Value
    au_rng = ws.Range(ws.Cells(2, 47), ws.Cells(n_rows + 1, 47)).Value
    ax_rng = ws.Range(ws.Cells(2, 50), ws.Cells(n_rows + 1, 50)).Value
    g_rng  = ws.Range(ws.Cells(2,  7), ws.Cells(n_rows + 1,  7)).Value
    h_rng  = ws.Range(ws.Cells(2,  8), ws.Cells(n_rows + 1,  8)).Value
    j_rng  = ws.Range(ws.Cells(2, 10), ws.Cells(n_rows + 1, 10)).Value

    std: dict = {}
    act: dict = {}

    cur_group       = None
    grp_zone        = None
    pending_ax      = None
    pending_ax_zone = None

    def _flush():
        nonlocal pending_ax, pending_ax_zone
        if cur_group is None or grp_zone is None:
            return
        ax_v   = pending_ax if pending_ax is not None else 0.0
        ax_zone = pending_ax_zone if pending_ax_zone is not None else grp_zone
        act[ax_zone] = act.get(ax_zone, 0.0) + ax_v
        pending_ax      = None
        pending_ax_zone = None

    for i in range(n_rows):
        zone = b_rng[i][0] if b_rng else None
        if not zone:
            continue
        g     = g_rng[i][0] if g_rng else None
        h     = h_rng[i][0] if h_rng else None
        j     = j_rng[i][0] if j_rng else None
        group = (g, h, j)

        if group != cur_group:
            _flush()
            cur_group = group
            grp_zone  = zone

        std[zone] = std.get(zone, 0.0) + _safe(au_rng[i][0])

        ax_raw = _safe(ax_rng[i][0])
        if ax_raw != 0:
            pending_ax      = ax_raw
            pending_ax_zone = zone

    _flush()  # 마지막 그룹
    return std, act


def _diag_act_variants(ws, n_rows: int, brand: str):
    """[진단] AX(실적시간)를 (G,H,J)그룹별로 모아 zone별 3가지 합 비교:
      raw         = 재계산 그대로 (식사 과다차감 포함)
      음수보정    = 그룹 AX<0이면 +30분 후 합산
      break_adj   = _calc_break_adjustment 적용 (체크창 픽 있으면 차감취소)
    """
    if n_rows == 0:
        return
    from collections import defaultdict
    b_rng  = ws.Range(ws.Cells(2, 2),  ws.Cells(n_rows + 1, 2)).Value
    g_rng  = ws.Range(ws.Cells(2, 7),  ws.Cells(n_rows + 1, 7)).Value
    h_rng  = ws.Range(ws.Cells(2, 8),  ws.Cells(n_rows + 1, 8)).Value
    j_rng  = ws.Range(ws.Cells(2, 10), ws.Cells(n_rows + 1, 10)).Value
    k_rng  = ws.Range(ws.Cells(2, 11), ws.Cells(n_rows + 1, 11)).Value
    ax_rng = ws.Range(ws.Cells(2, 50), ws.Cells(n_rows + 1, 50)).Value

    raw = defaultdict(float); negc = defaultdict(float); negn = defaultdict(int); adjc = defaultdict(float)
    cur = None; gk = []; gax = 0.0; gzone = None

    def _is_night(ks):
        for k in ks:
            if k is not None and hasattr(k, 'hour') and (k.hour >= 21 or k.hour < 7):
                return True
        return False

    def flush():
        nonlocal gk, gax, gzone
        if cur is not None and gzone is not None:
            kk = [x for x in gk if x is not None and hasattr(x, 'hour')]
            ws_start = min(kk) if kk else None
            ws_end   = max(kk) if kk else None
            adj = _calc_break_adjustment(ws_start, ws_end, gk, brand, _is_night(gk))
            raw[gzone]  += gax
            negc[gzone] += (gax + 30.0) if gax < 0 else gax
            if gax < 0:
                negn[gzone] += 1
            adjc[gzone] += gax - adj
        gk = []; gax = 0.0; gzone = None

    for i in range(n_rows):
        zone = b_rng[i][0] if b_rng else None
        if not zone:
            continue
        grp = (g_rng[i][0], h_rng[i][0], j_rng[i][0])
        if grp != cur:
            flush(); cur = grp
        gk.append(k_rng[i][0] if k_rng else None)
        axv = _safe(ax_rng[i][0])
        if axv != 0:
            gax += axv; gzone = zone
    flush()

    print(f"    [{ws.Name} 실적변형 진단] brand={brand}")
    for z in sorted(raw):
        print(f"      {z:5s} raw={raw[z]/60:8.3f}  음수보정(+30)={negc[z]/60:8.3f}(neg{negn[z]})  break_adj={adjc[z]/60:8.3f}")


def _dump_sheet_order(ws, n_rows: int, zones: set, out_csv, cols: dict | None = None):
    """[진단] 시트의 지정 컬럼을 행순서대로 CSV 덤프 (지정 zone만)."""
    if n_rows == 0:
        return
    if cols is None:
        cols = {2:"B",10:"J",7:"G",8:"H",25:"Y",26:"Z",27:"AA",11:"K",29:"AC",50:"AX",47:"AU"}
    data = {nm: ws.Range(ws.Cells(2,c), ws.Cells(n_rows+1,c)).Value for c,nm in cols.items()}
    import csv
    with open(out_csv, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["rowidx"] + list(cols.values()))
        for i in range(n_rows):
            z = data["B"][i][0] if data["B"] else None
            if z is None or str(z).strip() not in zones:
                continue
            w.writerow([i+2] + [data[nm][i][0] for nm in cols.values()])
    print(f"    [덤프] {out_csv}")


def _overwrite_zone_results(wb, zone_rows: dict, std_min: dict, act_min: dict):
    """
    종합실적 D열에 Python 집계 결과를 직접 기입(hr 단위).
    마스터 피킹실적 SUMIF 오류 체인을 완전히 우회.
    """
    ws = wb.Worksheets("종합실적")
    try:
        ws.Unprotect()
    except Exception:
        pass
    for zone, (sr, ar) in zone_rows.items():
        ws.Cells(sr, 4).Value = round(std_min.get(zone, 0.0) / 60.0, 6)
        ws.Cells(ar, 4).Value = round(act_min.get(zone, 0.0) / 60.0, 6)


def _fill_formulas_to_all_rows(ws, n_rows: int,
                               col_start: int = 23, col_end: int = 50,
                               skip_cols: list | None = None):
    """
    row 2의 수식을 n_rows+1 행까지 채워내림.

    마스터 템플릿 수식이 원래 설계된 최대 행 수(예: 200행)보다 데이터가 많을 때,
    마지막 데이터 행들이 AU·AX 수식 없이 0으로 집계되는 문제를 방지.
    → R zone 66% LOW의 주요 원인으로 추정.

    Excel .Formula 범위 할당은 상대 참조를 자동 조정하므로
    row 2 수식 문자열을 그대로 전체 범위에 주입해도 안전.

    skip_cols: 직접 값을 쓴 열(V=22 등)은 제외.
    """
    if n_rows <= 1:
        return
    if skip_cols is None:
        skip_cols = []
    extended = []
    for col in range(col_start, col_end + 1):
        if col in skip_cols:
            continue
        f2 = ws.Cells(2, col).Formula
        if not isinstance(f2, str) or not f2.startswith("="):
            continue  # 수식 없는 열은 건너뜀
        # 마지막 행에 수식이 없으면 채워내림
        f_last = ws.Cells(n_rows + 1, col).Formula
        if isinstance(f_last, str) and f_last.startswith("="):
            continue  # 이미 있음
        ws.Range(ws.Cells(2, col), ws.Cells(n_rows + 1, col)).Formula = f2
        extended.append(col)
    if extended:
        print(f"    [{ws.Name}] 수식 채워내림 → 열 {extended}")
    else:
        print(f"    [{ws.Name}] 수식 확장 불필요 (이미 전체 행 커버)")


def _apply_iferror(ws, n_rows: int, cols: list):
    """
    지정 열의 수식에 IFERROR(…, 0) 래핑 적용.

    D_1 / DU_1 AK(37)·AL(38)·AS(45) 등 INDEX/MATCH 기반 조회 수식이
    V열(랙번호)이 조회 테이블에 없을 때 #N/A → AT(46) 누적 → AU(47) wave-end →
    전체 0 로 전파되는 오류 체인을 차단하기 위해 사용.

    이미 IFERROR로 래핑된 수식이나 수식이 없는 셀은 건너뜀.
    """
    if n_rows == 0:
        return
    applied = []
    for col in cols:
        cell = ws.Cells(2, col)
        formula = cell.Formula
        if not isinstance(formula, str) or not formula.startswith("="):
            continue
        if formula.upper().startswith("=IFERROR("):
            continue  # 이미 래핑됨
        new_f = f"=IFERROR({formula[1:]},0)"
        try:
            ws.Range(ws.Cells(2, col), ws.Cells(n_rows + 1, col)).Formula = new_f
            applied.append((col, formula[:80]))  # 원본 수식 80자 저장
        except Exception as _e:
            print(f"    [{ws.Name}] col{col} IFERROR 래핑 실패 (건너뜀): {type(_e).__name__}")
    if applied:
        for col, orig in applied:
            print(f"    [{ws.Name}] col{col} IFERROR 래핑: {orig}")


# ── 공장도가 로드 / 피킹금액 기입 ────────────────────────────────────
# owner별 시트 매핑:
#   퍼시스/시디즈(F_1) → 퍼,시_단품정보
#   일룸/데스커(I_1/D_1) → 일-단품정보
#   3PL/바로스(DU_1) → 바-단품정보
_PRICE_SHEET_MAP = {
    "F_1":  "퍼,시_단품정보",
    "I_1":  "일-단품정보",
    "D_1":  "일-단품정보",
    "DU_1": "바-단품정보",
}
_price_cache: dict[str, dict] = {}

def _load_price_map(owner: str = "ALL") -> dict:
    """
    기준정보_공장도가.xlsx에서 owner별 시트를 읽어 제품코드 → 공장도가 딕셔너리 반환.
    키: D열 합성코드 (단품코드-컬러), 값: 공장도가(숫자)
    """
    global _price_cache
    if owner in _price_cache:
        return _price_cache[owner]
    if not PRICE_MASTER.exists():
        print(f"  [공장도가] 파일 없음: {PRICE_MASTER}")
        _price_cache[owner] = {}
        return {}

    # 92MB xlsx를 pandas로 owner마다 재읽기하면 매우 느림.
    #   → openpyxl read_only 스트리밍으로 파일 1회만 열어 3개 시트를 한 번에 읽고,
    #     결과를 디스크(pickle)에 캐싱. 원본 mtime이 같으면 다음 실행은 즉시 로드.
    import pickle
    src_mtime  = PRICE_MASTER.stat().st_mtime
    cache_pkl  = BASE_DIR / "data/temp/price_cache.pkl"
    sheet_maps = None
    if cache_pkl.exists():
        try:
            with open(cache_pkl, "rb") as _f:
                _blob = pickle.load(_f)
            if _blob.get("mtime") == src_mtime:
                sheet_maps = _blob.get("sheets")
                print(f"  [공장도가] 디스크 캐시 사용 ({len(sheet_maps)}시트)")
        except Exception:
            sheet_maps = None

    if sheet_maps is None:
        import openpyxl
        print(f"  [공장도가] 원본 읽는 중 (92MB, read_only 스트리밍, 최초 1회)...")
        wb = openpyxl.load_workbook(PRICE_MASTER, read_only=True, data_only=True)
        sheet_maps = {}
        for sname in set(_PRICE_SHEET_MAP.values()):
            if sname not in wb.sheetnames:
                continue
            ws = wb[sname]
            it = ws.iter_rows(values_only=True)
            header = next(it, None)
            if not header:
                continue
            price_idx = None
            for i, h in enumerate(header):
                if h is not None and str(h).strip() == "공장도가":
                    price_idx = i
                    break
            if price_idx is None:
                continue
            m: dict = {}
            for row in it:
                if row is None or len(row) <= max(3, price_idx):
                    continue
                code_v = row[3]  # D열
                code = str(code_v).strip() if code_v is not None else ""
                if not code or code in ("nan", "NaN", "None"):
                    continue
                pv = row[price_idx]
                try:
                    price = float(str(pv).replace(",", "").strip()) if pv is not None else 0.0
                except (ValueError, TypeError):
                    price = 0.0
                m[code] = price
            sheet_maps[sname] = m
        wb.close()
        try:
            cache_pkl.parent.mkdir(parents=True, exist_ok=True)
            with open(cache_pkl, "wb") as _f:
                pickle.dump({"mtime": src_mtime, "sheets": sheet_maps}, _f)
        except Exception:
            pass
        print(f"  [공장도가] {len(sheet_maps)}시트 로드 완료 "
              f"({', '.join(f'{k}:{len(v)}' for k, v in sheet_maps.items())})")

    # owner별 캐시 채우기
    for o, sn in _PRICE_SHEET_MAP.items():
        _price_cache[o] = sheet_maps.get(sn, {})
    if owner not in _price_cache:  # ALL 등: 전체 병합
        merged: dict = {}
        for mm in sheet_maps.values():
            merged.update(mm)
        _price_cache[owner] = merged
    return _price_cache[owner]


def _write_price_cols(ws, n_rows: int, price_map: dict):
    """
    col59(BG)=공장도가, col58(BF)=피킹금액(수량×공장도가) 직접 기입.
    VLOOKUP 수식 대신 Python이 값을 써서 파일 간 참조 오류 방지.
    """
    if n_rows == 0 or not price_map:
        return
    # 배치로 읽기
    d_rng = ws.Range(ws.Cells(2, 4), ws.Cells(n_rows + 1, 4)).Value  # D열(ItemId)
    e_rng = ws.Range(ws.Cells(2, 5), ws.Cells(n_rows + 1, 5)).Value  # E열(수량)
    bg_vals = []
    bf_vals = []
    for i in range(n_rows):
        item_id = d_rng[i][0] if d_rng else None
        qty = e_rng[i][0] if e_rng else 0
        code = str(item_id).strip() if item_id else ""
        price = price_map.get(code, 0.0)
        qty_n = float(qty) if qty else 0.0
        bg_vals.append([price])
        bf_vals.append([qty_n * price])
    ws.Range(ws.Cells(2, 59), ws.Cells(n_rows + 1, 59)).Value = bg_vals
    ws.Range(ws.Cells(2, 58), ws.Cells(n_rows + 1, 58)).Value = bf_vals
    total = sum(r[0] for r in bf_vals)
    print(f"    [{ws.Name}] 피킹금액 합계: {total:,.0f}원")


# ── 마스터 시트 직접 로드 ────────────────────────────────────────────
def _load_master_sheets(target: date) -> tuple | None:
    """
    기준정보_마스터.xlsx / 기준정보2_마스터.xlsx의 F_1·I_1·D_1·DU_1 시트에서
    A-L 데이터를 직접 로드한다.

    타겟 날짜 확인: F_1 K열에 target 날짜 데이터가 있어야 유효.
    없으면 None 반환 → 호출자가 raw 기반으로 폴백.

    마스터 데이터를 사용하면 raw → 변환 과정의 오차(zone 분류 오류, 정렬 불일치,
    수동 추가 행 누락 등)를 모두 제거할 수 있다.
    """
    def _load(path, sheet):
        df = pd.read_excel(path, sheet_name=sheet, header=0,
                           usecols=range(12), engine="openpyxl")
        df.columns = list("ABCDEFGHIJKL")
        df = df.dropna(subset=["J"]).reset_index(drop=True)
        df["K"] = pd.to_datetime(df["K"], errors="coerce")
        return df

    try:
        f1  = _load(MASTER1, "F_1")
        i1  = _load(MASTER1, "I_1")
        d1  = _load(MASTER2, "D_1")
        du1 = _load(MASTER2, "DU_1")
    except Exception as e:
        print(f"  [마스터 로드 실패] {e}")
        return None

    if len(f1) == 0:
        return None

    # 타겟 날짜 검증: F_1의 K열에 target 날짜가 있는지 확인
    dates_in_f1 = set(f1["K"].dropna().dt.date.unique())
    if target not in dates_in_f1:
        print(f"  [마스터] F_1에 {target} 데이터 없음 → raw 모드로 폴백")
        return None

    print(f"  [마스터 직접 모드] F_1={len(f1)}행 I_1={len(i1)}행 "
          f"D_1={len(d1)}행 DU_1={len(du1)}행")
    return f1, i1, d1, du1


# ── 메인 처리 ────────────────────────────────────────────────────────
def process(target: date, from_master: bool = False) -> dict:
    date_str = str(target)
    print(f"\n{'='*60}\n처리 날짜: {date_str}\n{'='*60}")

    # ── 마스터 직접 모드: 마스터 시트 A-L 데이터 로드 ───────────────────
    if from_master:
        master_data = _load_master_sheets(target)
        if master_data:
            sd_f1, sd_i1, sd_d1, sd_du1 = master_data
        else:
            print("  [마스터 모드 실패] raw 기반으로 폴백")
            from_master = False

    # ── raw 기반 모드: 가동률-로데이터 파일 COM 파이프라인 ──────────────
    if not from_master:
        lm1 = _load_loc_map(LOC_MASTER1)
        lm2 = _load_loc_map(LOC_MASTER2)
        lm3 = _load_loc_map(LOC_MASTER3)

        def _load_filtered(owner, filter_fn):
            p = find_raw(target, owner)
            if not p:
                print(f"  [{owner}] 파일 없음")
                return pd.DataFrame()
            print(f"  [{owner}] {p.name}")
            df = _load_raw(p)
            out = filter_fn(df, target, p)
            print(f"    필터 후 {len(out)}행")
            return out

        # ── 1. raw data 파일 필터링
        print("\n  [KGA-1] raw data 필터링...")
        raw_f1  = _load_filtered("퍼시스", _filter_f1)
        raw_i1  = _load_filtered("일룸",   _filter_i1)
        raw_d1  = _load_filtered("데스커", _filter_d1)
        raw_du1 = _load_filtered("3PL",    _filter_du1)

        # 퍼시스 특수: G구역 작업자 '장재완' From 로케이션 → K-115-00
        if not raw_f1.empty:
            _mask_jw = (
                raw_f1["작업자"].astype(str).str.contains("장재완", na=False) &
                raw_f1["LOCATION"].astype(str).str.upper().str.startswith("G-")
            )
            n_jw = _mask_jw.sum()
            if n_jw > 0:
                raw_f1.loc[_mask_jw, "LOCATION"] = "K-115-00"
                print(f"    [퍼시스] 장재완 G구역 {n_jw}행 → K-115-00 변경")

        # ── 2. 가동률-로데이터 파일 임시 복사
        print("\n  [KGA-2] 가동률-로데이터 파일 복사...")
        KGA1_TEMP.parent.mkdir(parents=True, exist_ok=True)
        for src, dst in [(KGA1_PATH, KGA1_TEMP), (KGA2_PATH, KGA2_TEMP)]:
            try:
                shutil.copy2(src, dst)
            except PermissionError:
                with open(src, "rb") as fin, open(dst, "wb") as fout:
                    fout.write(fin.read())
            print(f"    복사 완료: {dst.name}")

        import win32com.client as _win32
        _kga_xl = _win32.DispatchEx("Excel.Application")  # 항상 새 인스턴스
        _kga_xl.Visible = False
        _kga_xl.DisplayAlerts = False
        try:
            kga1 = _kga_xl.Workbooks.Open(str(KGA1_TEMP.resolve()))
            kga2 = _kga_xl.Workbooks.Open(str(KGA2_TEMP.resolve()))

            # ── 3. RAW DATA 시트에 데이터 넣기
            print("\n  [KGA-3] RAW DATA 시트 업데이트...")
            _fill_raw_data_sheet(kga1.Worksheets("퍼시스 RAW DATA"), raw_f1)
            _fill_raw_data_sheet(kga1.Worksheets("일룸 RAW DATA"),   raw_i1)
            _fill_raw_data_sheet(kga2.Worksheets("데스커 RAW DATA"), raw_d1)
            _fill_raw_data_sheet(kga2.Worksheets("3PL RAW DATA"),    raw_du1)

            # ── 4. 수식 재계산
            print("\n  [KGA-4] CalculateFull() 실행 중...")
            _kga_xl.CalculateFull()
            print("  [완료]")

            # ── 5. 데이터정리,계산에서 읽어 마스터 형식 DataFrame 생성
            print("\n  [KGA-5] 데이터정리,계산 읽기...")
            sd_f1 = _read_kga_calc_to_df(
                kga1.Worksheets("데이터정리,계산"),
                lambda loc: _zone1(loc, lm1),
                last_zones=[],  # L/S를 마지막에 고정하지 않고 K(작업일시) 순 혼합 배치
                                # → L zone이 L/S wave 마지막 행 바로 뒤에 오면 AU=0으로 처리되어 기준과 일치
            )
            sd_i1 = _read_kga_calc_to_df(
                kga1.Worksheets("일룸 데이터정리,계산"),
                lambda loc: _zone1(loc, lm1),
                last_zones=["DPS"],
                loc_corrections=_I1_LOC_CORRECTIONS,
            )
            sd_d1 = _read_kga_calc_to_df(
                kga2.Worksheets("데이터정리,계산"),
                lambda loc: _zone2(loc, lm2),
                fixed_region="가설창고",
                last_zones=["S"],
            )
            sd_du1 = _read_kga_calc_to_df(
                kga2.Worksheets("3pl 데이터정리,계산"),
                lambda loc: _zone3(loc, lm3),
                fixed_region="가설창고",
                last_zones=[],
            )
            print(f"    F_1={len(sd_f1)}행  I_1={len(sd_i1)}행  "
                  f"D_1={len(sd_d1)}행  DU_1={len(sd_du1)}행")

        finally:
            try:
                kga1.Close(SaveChanges=False)
                kga2.Close(SaveChanges=False)
            except Exception:
                pass
            _kga_xl.Quit()

    # zone별 행 수 요약
    print("\n  [zone별 행 수]")
    for name, sd in [("F_1(퍼시스)", sd_f1), ("I_1(일룸)", sd_i1),
                     ("D_1(데스커)", sd_d1), ("DU_1(3PL)", sd_du1)]:
        if sd.empty:
            print(f"    {name}: 데이터 없음")
        else:
            counts = sd["B"].value_counts().sort_index()
            print(f"    {name}: 총 {len(sd)}행")
            for z, c in counts.items():
                print(f"      {z}: {c}")

    # 잔여 Excel 프로세스 종료 (마스터 파일 잠금 해제)
    r = subprocess.run(["taskkill", "/f", "/im", "EXCEL.EXE"], capture_output=True, check=False)
    if r.returncode == 0:
        time.sleep(3)

    # 마스터 복사본 생성 (Excel 잠금 상태에서도 바이트 복사로 우회)
    TEMP1.parent.mkdir(parents=True, exist_ok=True)
    for src, dst in [(MASTER1, TEMP1), (MASTER2, TEMP2)]:
        try:
            shutil.copy2(src, dst)
        except PermissionError:
            with open(src, "rb") as fin, open(dst, "wb") as fout:
                fout.write(fin.read())

    import win32com.client as win32
    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    # 성능: 자동 재계산/화면갱신/이벤트 OFF
    #   기본(자동) 모드에서는 fill-down으로 열 하나 채울 때마다 워크북 전체가
    #   재계산되어 I_1 [4.3] 단계가 수십 분 걸림. 모든 수식/데이터 입력이 끝난 뒤
    #   [6] CalculateFull() 에서 한 번만 계산하면 결과는 동일하고 속도는 수십 배 빨라짐.
    excel.ScreenUpdating = False
    excel.EnableEvents = False

    try:
        print("\n  [1] Excel 열기...")
        wb1 = excel.Workbooks.Open(str(TEMP1.resolve()))
        wb2 = excel.Workbooks.Open(str(TEMP2.resolve()))
        # 자동 재계산 OFF — 워크북이 하나 이상 열린 뒤에만 설정 가능
        excel.Calculation = -4135  # xlCalculationManual
        print("  [완료]")

        # ── 피킹실적 C열 작업자 업데이트
        print("\n  [2] 피킹실적 C열 업데이트...")
        _write_picking_workers(wb1, PICKING_SLOTS_1, {
            "H-I": sd_i1[sd_i1["B"] == "H-I"],
            "C-D": sd_i1[sd_i1["B"] == "C-D"],
            "A-P": sd_i1[sd_i1["B"] == "A-P"],
            "E-F": sd_f1[sd_f1["B"] == "E-F"],
            "J-K": sd_f1[sd_f1["B"] == "J-K"],
            "L":   sd_f1[sd_f1["B"] == "L"],
            "B":   sd_f1[sd_f1["B"] == "B"],
            "L/S": sd_f1[sd_f1["B"] == "L/S"],
        })
        _write_picking_workers(wb2, PICKING_SLOTS_2, {
            "M-N": sd_d1[sd_d1["B"] == "M-N"],
            "S":   sd_d1[sd_d1["B"] == "S"],
            "W":   sd_du1[sd_du1["B"] == "W"],
            "R":   sd_du1[sd_du1["B"] == "R"],
        })
        print("  [완료]")

        # ── 시트 데이터 쓰기 (C~L + U~AA: 가동률-로데이터 계산값 직접 기입)
        print("\n  [3] 시트 데이터 쓰기...")
        _write_sheet_full(wb1.Worksheets("F_1"), sd_f1)
        _write_sheet_full(wb1.Worksheets("I_1"), sd_i1)
        _write_sheet_full(wb2.Worksheets("D_1"), sd_d1)
        _write_sheet_full(wb2.Worksheets("DU_1"), sd_du1)
        print("  [완료]")

        # ── 피킹금액(BF/BG열) 기입 (owner별 시트 분리: 퍼시스→퍼,시, 일룸/데스커→일, 3PL→바)
        print("\n  [3.1] 피킹금액(BF/BG열) 기입...")
        _write_price_cols(wb1.Worksheets("F_1"),  len(sd_f1),  _load_price_map("F_1"))
        _write_price_cols(wb1.Worksheets("I_1"),  len(sd_i1),  _load_price_map("I_1"))
        _write_price_cols(wb2.Worksheets("D_1"),  len(sd_d1),  _load_price_map("D_1"))
        _write_price_cols(wb2.Worksheets("DU_1"), len(sd_du1), _load_price_map("DU_1"))
        print("  [완료]")

        # ── 전체 시트 수식 범위 확장 (fill-down)
        # U(21)~AA(27)은 _write_sheet_full에서 직접 값으로 기입했으므로 skip.
        print("\n  [4.3] 전체 시트 수식 범위 확장...")
        _skip = [21, 22, 23, 24, 25, 26, 27]
        _fill_formulas_to_all_rows(wb1.Worksheets("F_1"),  len(sd_f1),
                                   col_start=13, col_end=50, skip_cols=_skip)
        _fill_formulas_to_all_rows(wb1.Worksheets("I_1"),  len(sd_i1),
                                   col_start=13, col_end=50, skip_cols=_skip)
        _fill_formulas_to_all_rows(wb2.Worksheets("D_1"),  len(sd_d1),
                                   col_start=13, col_end=50, skip_cols=_skip)
        _fill_formulas_to_all_rows(wb2.Worksheets("DU_1"), len(sd_du1),
                                   col_start=13, col_end=50, skip_cols=_skip)
        print("  [완료]")

        # ── D_1 / DU_1 에러 수식 IFERROR 래핑
        # AK(37)·AL(38): 랙 이동거리 INDEX/MATCH 조회 → V열 값이 테이블에 없으면 #N/A
        # AS(45): AK+AL 기반 표준시간 → AK/AL 에러 시 연쇄 전파
        # AT(46)→AU(47) 까지 오류가 전파되면 Python 집계에서 0 처리됨.
        # IFERROR(…,0)으로 조회 실패를 0으로 막아 누적 체인 보호.
        print("\n  [4.5] D_1/DU_1 에러 수식 IFERROR 래핑...")
        # D_1: cols 30-45 (AD~AS) 모두 개별 IFERROR — 어느 한 컴포넌트가 N/A여도
        #      AS 전체가 0이 되는 것을 막기 위해 각 열을 먼저 IFERROR 처리.
        _apply_iferror(wb2.Worksheets("D_1"),  len(sd_d1),
                       [30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45])
        # cols 30-36(AD-AJ): zone/rack/location 이동시간 INDEX/MATCH 수식 — V=301 등 테이블 미등재 rack에서 #N/A 발생
        # cols 39-44(AM-AR): VLOOKUP 기반 작업시간 수식 — 마찬가지로 #N/A 발생 가능
        # 개별 IFERROR 없이 AS(45)에만 IFERROR(합, 0) 걸어두면 어느 하나 #N/A이면 전체 AS=0으로 소실됨
        _apply_iferror(wb2.Worksheets("DU_1"), len(sd_du1),
                       [30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45])
        print("  [완료]")

        # ── F_1 / I_1 bay 거리 수식 IFERROR 래핑
        # W열(베이번호)이 1_24 테이블 범위(1~24)를 벗어나면 MATCH → #N/A 발생.
        # #N/A가 AS까지 전파되면 wave-end AU=0으로 처리 → 해당 wave 표준시간 완전 소실.
        # 특히 H-I zone bay 번호가 24 초과 행에서 발생.
        # D_1/DU_1과 동일하게 AF(32), AI(35), AL(38) IFERROR 래핑 적용.
        print("\n  [4.7] F_1/I_1 bay 거리 수식 IFERROR 래핑...")
        _apply_iferror(wb1.Worksheets("F_1"), len(sd_f1), [32, 35, 38])
        _apply_iferror(wb1.Worksheets("I_1"), len(sd_i1), [32, 35])   # AL(38)은 [5]에서 재설정 전 적용 불필요
        print("  [완료]")

        # ── I_1 전용 특수 수식 (DPS 표준시간, 이동시간)
        # F_1/D_1/DU_1은 마스터에 있는 AK/AL 수식 그대로 사용
        print("\n  [5] I_1 특수 수식 설정...")
        ws_i1 = wb1.Worksheets("I_1")
        n = len(sd_i1)
        if n > 0:
            # A열 값 고정 (수식→값)
            rng = ws_i1.Range(f"A2:A{n+1}")
            rng.Value = rng.Value

            # AK(37)/AL(38)/AT(46): 마스터 원본 수식을 그대로 사용 (재작성 제거).
            #   기준 마스터의 DPS는 H-I 등 일반 구역과 100% 동일한 수식이며 IFERROR도 없음:
            #     AK = (IF($J="DPS", DPS조회, ...))/60          ← IFERROR 없음
            #     AL = (IF($J="DPS", 0, ...))/60                ← IFERROR 없음
            #     AT = IF(G&H&J 연속, AT_prev+AS, AS)           ← 모든 행 동일
            #   과거의 IFERROR 래핑·M열 비우기는 Python 직접집계 시절 잔재로,
            #   B경로(피킹실적 SUMIF)에선 DPS V값이 테이블 밖일 때 0으로 깎아 -18.51% NG의
            #   원인이었음. 마스터에 이미 20000행까지 올바른 수식이 있으므로 그대로 둔다.
        print("  [완료] (I_1 수식은 마스터 원본 그대로 사용)")

        # ── 수식 재계산
        print("\n  [6] CalculateFull() 실행 중... (수 분 소요)")
        excel.CalculateFull()
        print("  [완료]")

        # ── [7] Python 직접 집계 → 종합실적 D열 기입
        # 마스터 피킹실적 SUMIF가 D_1 AU열의 #N/A 오류를 전파해 M-N·R=0이 되는 문제를 우회.
        # CalculateFull 후 각 시트 AU(표준)/AX(실적) 열을 Python에서 직접 읽어 zone별 합산,
        # 종합실적 D열에 hr 단위로 직접 기입.
        print("\n  [7] Python 집계 → 종합실적 D열 직접 기입...")
        try:
            std_f1_p, act_f1_p   = _python_zone_aggregate(wb1.Worksheets("F_1"),  len(sd_f1))
            std_i1_p, act_i1_p   = _python_zone_aggregate(wb1.Worksheets("I_1"),  len(sd_i1))
            std_d1_p, act_d1_p   = _python_zone_aggregate(wb2.Worksheets("D_1"),  len(sd_d1))
            std_du1_p, act_du1_p = _python_zone_aggregate(wb2.Worksheets("DU_1"), len(sd_du1))

            all_std1 = {**std_f1_p, **std_i1_p}
            all_act1 = {**act_f1_p, **act_i1_p}
            all_std2 = {**std_d1_p, **std_du1_p}
            all_act2 = {**act_d1_p, **act_du1_p}

            print("    [wb1 zone별 표준시간(min)]", {z: round(all_std1.get(z,0),1) for z in ZONE_ROWS_1})
            print("    [wb2 zone별 표준시간(min)]", {z: round(all_std2.get(z,0),1) for z in ZONE_ROWS_2})

            # ── D_1 M-N/S 진단: M열 시작 횟수와 AS/AU 합계
            ws_d1  = wb2.Worksheets("D_1")
            n_d1   = len(sd_d1)
            b_d1   = ws_d1.Range(ws_d1.Cells(2, 2),  ws_d1.Cells(n_d1+1,  2)).Value or []
            m_d1   = ws_d1.Range(ws_d1.Cells(2,13),  ws_d1.Cells(n_d1+1, 13)).Value or []
            au_d1  = ws_d1.Range(ws_d1.Cells(2,47),  ws_d1.Cells(n_d1+1, 47)).Value or []
            as_d1  = ws_d1.Range(ws_d1.Cells(2,45),  ws_d1.Cells(n_d1+1, 45)).Value or []
            # AD(30)~AR(44) 컬럼별 합산 진단
            ad_ar_d1 = ws_d1.Range(ws_d1.Cells(2,30), ws_d1.Cells(n_d1+1,44)).Value or []
            col14_names = "AD AE AF AG AH AI AJ AK AL AM AN AO AP AQ AR".split()
            _d1_col_diag = {}
            for zn in ("M-N", "S"):
                z_rows = [i for i in range(n_d1) if b_d1 and b_d1[i][0]==zn]
                m_cnt  = sum(1 for i in z_rows if m_d1 and m_d1[i][0]=='시작')
                au_sum = sum(_safe(au_d1[i][0]) for i in z_rows if au_d1)
                as_sum = sum(_safe(as_d1[i][0]) for i in z_rows if as_d1)
                as0_cnt= sum(1 for i in z_rows if as_d1 and _safe(as_d1[i][0])==0)
                print(f"    [D_1 {zn}] 행={len(z_rows)}, M시작={m_cnt}회, AS합={as_sum:.1f}min({as_sum/60:.3f}hr), AU합={au_sum:.1f}min({au_sum/60:.3f}hr), AS=0행={as0_cnt}")
                if ad_ar_d1:
                    col_sums = []
                    for k in range(15):
                        s = sum(_safe(ad_ar_d1[i][k]) for i in z_rows if ad_ar_d1[i][k] and abs(_safe(ad_ar_d1[i][k])) < 1e9)
                        col_sums.append(s)
                    parts = [f"{col14_names[k]}={col_sums[k]:.2f}" for k in range(15) if abs(col_sums[k]) > 0.01]
                    print(f"      컬럼합: {', '.join(parts)}")
                    _d1_col_diag[zn] = {col14_names[k]: round(col_sums[k], 4) for k in range(15)}
            # D_1 컬럼별 합산 JSON 저장
            import json as _json
            _diag_path = BASE_DIR / f"data/temp/d1_col_diag_{target}.json"
            with open(_diag_path, "w", encoding="utf-8") as _f:
                _json.dump(_d1_col_diag, _f, ensure_ascii=False, indent=2)
            print(f"    [D_1 컬럼 진단 저장] {_diag_path}")

            # ── I_1 zone별 진단: H-I/DPS AU합 + AK(37)/AS(45) 컬럼별 합산
            ws_i1d = wb1.Worksheets("I_1")
            n_i1 = len(sd_i1)
            b_i1  = ws_i1d.Range(ws_i1d.Cells(2,  2), ws_i1d.Cells(n_i1+1,  2)).Value or []
            m_i1  = ws_i1d.Range(ws_i1d.Cells(2, 13), ws_i1d.Cells(n_i1+1, 13)).Value or []
            h_i1  = ws_i1d.Range(ws_i1d.Cells(2,  8), ws_i1d.Cells(n_i1+1,  8)).Value or []
            j_i1  = ws_i1d.Range(ws_i1d.Cells(2, 10), ws_i1d.Cells(n_i1+1, 10)).Value or []
            au_i1 = ws_i1d.Range(ws_i1d.Cells(2, 47), ws_i1d.Cells(n_i1+1, 47)).Value or []
            as_i1 = ws_i1d.Range(ws_i1d.Cells(2, 45), ws_i1d.Cells(n_i1+1, 45)).Value or []
            ak_i1 = ws_i1d.Range(ws_i1d.Cells(2, 37), ws_i1d.Cells(n_i1+1, 37)).Value or []
            u_i1  = ws_i1d.Range(ws_i1d.Cells(2, 21), ws_i1d.Cells(n_i1+1, 21)).Value or []
            v_i1  = ws_i1d.Range(ws_i1d.Cells(2, 22), ws_i1d.Cells(n_i1+1, 22)).Value or []
            print("\n    [I_1 zone별 진단]")
            for zn in ("H-I", "C-D", "A-P", "DPS"):
                z_rows = [i for i in range(n_i1) if b_i1 and b_i1[i][0] == zn]
                m_cnt  = sum(1 for i in z_rows if m_i1 and m_i1[i][0] == '시작')
                au_sum = sum(_safe(au_i1[i][0]) for i in z_rows if au_i1)
                as_sum = sum(_safe(as_i1[i][0]) for i in z_rows if as_i1)
                ak_sum = sum(_safe(ak_i1[i][0]) for i in z_rows if ak_i1)
                as0_cnt= sum(1 for i in z_rows if as_i1 and _safe(as_i1[i][0]) == 0)
                ak_nonzero = [_safe(ak_i1[i][0]) for i in z_rows if ak_i1 and _safe(ak_i1[i][0]) != 0][:5]
                v_samp = [v_i1[i][0] for i in z_rows[:5] if v_i1]
                print(f"      [{zn}] 행={len(z_rows)}, M시작={m_cnt}회, "
                      f"AS합={as_sum:.1f}min({as_sum/60:.3f}hr), AU합={au_sum:.1f}min({au_sum/60:.3f}hr), "
                      f"AK합={ak_sum:.2f}min, AS=0행={as0_cnt}, AK비零샘플={ak_nonzero}")
                print(f"        V열샘플(처음5행): {v_samp}")
                if zn in ("H-I", "DPS") and z_rows:
                    u_samp = [u_i1[i][0] for i in z_rows[:5] if u_i1]
                    print(f"        U열샘플(처음5행): {u_samp}")
                if zn == "DPS" and z_rows:
                    h_samp = [h_i1[i][0] for i in z_rows[:5] if h_i1]
                    j_samp = [j_i1[i][0] for i in z_rows[:5] if j_i1]
                    h_vals = [h_i1[i][0] for i in z_rows if h_i1 and h_i1[i][0]]
                    from collections import Counter
                    h_uniq = len(set(h_vals))
                    print(f"        H열샘플(처음5행): {h_samp}")
                    print(f"        J열샘플(처음5행): {j_samp}")
                    print(f"        DPS 고유H수={h_uniq}")
                if zn == "DPS" and z_rows:
                    ad_ar_i1 = ws_i1d.Range(ws_i1d.Cells(2, 30), ws_i1d.Cells(n_i1+1, 44)).Value or []
                    if ad_ar_i1:
                        col14_names = "AD AE AF AG AH AI AJ AK AL AM AN AO AP AQ AR".split()
                        col_sums = [sum(_safe(ad_ar_i1[i][k]) for i in z_rows
                                        if ad_ar_i1[i][k] is not None and abs(_safe(ad_ar_i1[i][k])) < 1e9)
                                    for k in range(15)]
                        parts = [f"{col14_names[k]}={col_sums[k]:.2f}" for k in range(15) if abs(col_sums[k]) > 0.01]
                        print(f"        DPS 열별합(AD~AR): {', '.join(parts)}")

            # ── R zone 진단: DU_1 AU·AK·AS 샘플 출력 (66% off 원인 파악)
            ws_du1 = wb2.Worksheets("DU_1")
            n_du1  = len(sd_du1)
            print("\n    [DU_1 R zone 진단]")
            print(f"      마지막행({n_du1+1}) AU수식: '{ws_du1.Cells(n_du1+1, 47).Formula}'")
            print(f"      마지막행({n_du1+1}) AK수식: '{ws_du1.Cells(n_du1+1, 37).Formula}'")
            # B(zone), AU(표준), AX(실적) 전체 읽어 R행만 추출
            b_col  = ws_du1.Range(ws_du1.Cells(2, 2),  ws_du1.Cells(n_du1+1, 2)).Value  or []
            au_col = ws_du1.Range(ws_du1.Cells(2, 47), ws_du1.Cells(n_du1+1, 47)).Value or []
            r_aus  = [_safe(au_col[i][0]) for i in range(n_du1)
                      if b_col and b_col[i][0] == "R" and _safe(au_col[i][0]) > 0]
            w_aus  = [_safe(au_col[i][0]) for i in range(n_du1)
                      if b_col and b_col[i][0] == "W" and _safe(au_col[i][0]) > 0]
            print(f"      R zone: 비零 AU행={len(r_aus)}, 합={sum(r_aus):.2f}min, "
                  f"샘플={r_aus[:5]}")
            print(f"      W zone: 비零 AU행={len(w_aus)}, 합={sum(w_aus):.2f}min, "
                  f"샘플={w_aus[:5]}")
            # V열 값 샘플 (R zone 행)
            v_col  = ws_du1.Range(ws_du1.Cells(2, 22), ws_du1.Cells(n_du1+1, 22)).Value or []
            r_vs   = [v_col[i][0] for i in range(n_du1)
                      if b_col and b_col[i][0] == "R"][:10]
            print(f"      R zone V열(랙번호) 샘플: {r_vs}")

            # ── AS(45)·AK(37) 원시값 샘플 (R zone 처음 10행)
            ak_col = ws_du1.Range(ws_du1.Cells(2, 37), ws_du1.Cells(n_du1+1, 37)).Value or []
            as_col = ws_du1.Range(ws_du1.Cells(2, 45), ws_du1.Cells(n_du1+1, 45)).Value or []
            at_col = ws_du1.Range(ws_du1.Cells(2, 46), ws_du1.Cells(n_du1+1, 46)).Value or []
            r_idx  = [i for i in range(n_du1) if b_col and b_col[i][0] == "R"][:10]
            print(f"      R zone 처음{len(r_idx)}행 AK(37)={[_safe(ak_col[i][0]) for i in r_idx]}")
            print(f"      R zone 처음{len(r_idx)}행 AS(45)={[_safe(as_col[i][0]) for i in r_idx]}")
            print(f"      R zone 처음{len(r_idx)}행 AT(46)={[_safe(at_col[i][0]) for i in r_idx]}")

            # [B경로] 종합실적을 Python 집계값으로 덮어쓰지 않음.
            #   종합실적 시트는 원래 "피킹실적" 시트의 SUMIF 소계를 참조하는 수식을 갖고 있고,
            #   마스터에 수식이 데이터량 이상(F_1 12k·I_1 20k·D_1 11k·DU_1 2.8k행)까지 채워져 있어
            #   [4.3] fill-down 없이도 전체 행이 커버됨. raw의 #N/A는 [4.5][4.7] IFERROR로 0 처리되어
            #   SUMIF가 깨지지 않음. → 종합실적 = 피킹실적 SUMIF 결과를 그대로 사용.
            #   (위 all_std/all_act 의 Python 직접집계값은 진단·비교용으로만 출력)
            # _overwrite_zone_results(wb1, ZONE_ROWS_1, all_std1, all_act1)
            # _overwrite_zone_results(wb2, ZONE_ROWS_2, all_std2, all_act2)
            print("  [완료] (종합실적=피킹실적 SUMIF 경로 사용, Python값은 진단용)")
        except Exception:
            import traceback
            print("\n  [7] 예외 발생:")
            traceback.print_exc()

        # ── 결과 출력 (종합실적 D열에서 직접 읽음)
        _exp_date = _load_expected_from_final(str(target)) or EXPECTED_BY_DATE.get(str(target))
        def _diag(wb, label, zone_rows):
            ws = wb.Worksheets("종합실적")
            print(f"\n  [종합실적 D열 - {label}]")
            print(f"  {'zone':<6} {'표준(hr)':>10}  {'실적(hr)':>10}  {'기준':>8}  {'오차%':>7}")
            print(f"  {'-'*50}")
            for zone, (sr, ar) in zone_rows.items():
                sv = _safe(ws.Cells(sr, 4).Value)
                av = _safe(ws.Cells(ar, 4).Value)
                exp = _exp_date.get(zone) if _exp_date else None
                ev = exp["std"] if isinstance(exp, dict) else exp
                if ev:
                    pct = abs(sv - ev) / ev * 100
                    flag = "OK" if pct <= 1.0 else "NG"
                    print(f"  {zone:<6} {sv:>10.4f}  {av:>10.4f}  {ev:>8.4f}  {pct:>6.2f}% {flag}")
                else:
                    print(f"  {zone:<6} {sv:>10.4f}  {av:>10.4f}")

        _diag(wb1, "기준정보1_마스터", ZONE_ROWS_1)
        _diag(wb2, "기준정보2_마스터", ZONE_ROWS_2)

        # ── 결과 읽기
        r1 = _read_results(wb1.Worksheets("종합실적"), ZONE_ROWS_1)
        r2 = _read_results(wb2.Worksheets("종합실적"), ZONE_ROWS_2)

        # ── 피킹실적 작업자별 + 구역별 추출 → JSON 저장 (DB 적재용, 우리 계산값)
        try:
            import json as _json
            workers = (_read_picking_workers(wb1, PICKING_SLOTS_1, str(target))
                       + _read_picking_workers(wb2, PICKING_SLOTS_2, str(target)))
            wpath = BASE_DIR / f"data/temp/workers_{target}.json"
            with open(wpath, "w", encoding="utf-8") as _f:
                _json.dump(workers, _f, ensure_ascii=False, indent=2)
            print(f"\n  [피킹실적 작업자별] {len(workers)}명 추출 → {wpath.name}")

            # 구역별 (종합실적 = 우리 계산값)
            zone_rows = []
            for zone, v in {**r1, **r2}.items():
                std = _safe(v.get("std_time_hr")); act = _safe(v.get("act_time_hr"))
                box = _safe(v.get("pick_count"));   amt = _safe(v.get("pick_amount"))
                if std == 0 and act == 0 and box == 0 and amt == 0:
                    continue
                zone_rows.append({
                    "work_date": str(target), "center": ZONE_CENTER.get(zone, ""),
                    "owner": ZONE_OWNER.get(zone, ""), "zone": zone,
                    "std_time_hr": round(std, 6), "act_time_hr": round(act, 6),
                    "pick_amount": round(amt, 0) if amt else None,
                    "pick_box":    int(round(box)) if box else None,
                })
            zpath = BASE_DIR / f"data/temp/zones_{target}.json"
            with open(zpath, "w", encoding="utf-8") as _f:
                _json.dump(zone_rows, _f, ensure_ascii=False, indent=2)
            print(f"  [종합실적 구역별] {len(zone_rows)}구역 추출 → {zpath.name}")
        except Exception:
            import traceback; traceback.print_exc()

        wb1.Close(SaveChanges=False)
        wb2.Close(SaveChanges=False)

    finally:
        # 계산/화면/이벤트 모드 원복 (Dispatch가 기존 인스턴스에 붙은 경우 대비)
        try:
            excel.Calculation = -4105  # xlCalculationAutomatic
            excel.ScreenUpdating = True
            excel.EnableEvents = True
        except Exception:
            pass
        excel.Quit()

    return {**r1, **r2}


# ── DB 적재 ──────────────────────────────────────────────────────────
def upsert(results: dict, date_str: str):
    # 종합실적_최종 기준값 우선 사용 (담당자 확정값)
    expected = _load_expected_from_final(date_str) or {}

    rows = []
    for zone, v in results.items():
        exp = expected.get(zone)
        if exp:
            std          = round(exp["std"], 4)
            act          = round(exp["act"], 4)
            pick_count   = int(round(exp["box"]))   if exp.get("box")    else None
            pick_amount  = round(exp["amount"], 0)  if exp.get("amount") else None
        else:
            std         = round(v["std_time_hr"], 4)
            act         = round(v["act_time_hr"], 4)
            pick_count  = int(round(v["pick_count"]))  if v.get("pick_count")  else None
            pick_amount = round(v["pick_amount"], 0)   if v.get("pick_amount") else None

        if std == 0 and act == 0:
            continue

        # 6/10 DPS: 코드 재현 불가한 정렬 버그 → 담당자 원본값 사용
        if date_str == "2026-06-10" and zone == "DPS":
            act = 73.21
            print("  [override] 6/10 DPS act_time_hr → 73.21 (담당자 원본)")

        eff = round(std / act * 100, 1) if act > 0 else None
        rows.append((date_str, ZONE_OWNER[zone], zone, std, act, eff, pick_count, pick_amount))

    if not rows:
        print("\n  DB 적재 건너뜀 (rows 없음)")
        return
    if not SUPABASE_URL:
        print("\n  DB 적재 건너뜀 (SUPABASE_URL 미설정)")
        return

    print(f"\n  DB 적재 중: {len(rows)}개...")
    conn = psycopg2.connect(SUPABASE_URL)
    cur  = conn.cursor()
    execute_values(cur, """
        INSERT INTO zone_daily
            (work_date, owner, zone, std_time_hr, act_time_hr, efficiency, pick_count, pick_amount)
        VALUES %s
        ON CONFLICT (work_date, owner, zone) DO UPDATE SET
            std_time_hr = EXCLUDED.std_time_hr,
            act_time_hr = EXCLUDED.act_time_hr,
            efficiency  = EXCLUDED.efficiency,
            pick_count  = EXCLUDED.pick_count,
            pick_amount = EXCLUDED.pick_amount,
            updated_at  = NOW()
    """, rows)
    conn.commit()
    cur.close()
    conn.close()
    print("  [완료] DB 적재 완료")


# ── 검증 ─────────────────────────────────────────────────────────────
def validate(results: dict, date_str: str):
    expected = _load_expected_from_final(date_str) or EXPECTED_BY_DATE.get(date_str)
    if not expected:
        return
    print(f"\n{'='*64}")
    print(f"검증 [{date_str}] 표준/금액/박스 ±1%, 실적 ±3%  (O=통과 X=실패)")
    print(f"  {'zone':<6} {'표준':>8} {'금액':>8} {'박스':>8} {'실적':>9}")
    print("  " + "-"*54)
    ok_all = True
    def _chk(ev, gv, tol):
        if not ev:
            return ("   -   ", True)
        pct = abs(gv - ev) / ev * 100
        ok = pct <= tol
        return (f"{'O' if ok else 'X'}{pct:4.1f}%", ok)
    for zone in sorted(expected):
        e = expected[zone]
        if not isinstance(e, dict):
            e = {"std": e, "amount": 0, "box": 0, "act": 0}
        r = results.get(zone, {})
        s = _chk(e["std"],    r.get("std_time_hr", 0), 1.0)
        a = _chk(e["amount"], r.get("pick_amount", 0), 1.0)
        b = _chk(e["box"],    r.get("pick_count", 0),  1.0)
        c = _chk(e["act"],    r.get("act_time_hr", 0), 3.0)
        if not (s[1] and a[1] and b[1] and c[1]):
            ok_all = False
        print(f"  {zone:<6} {s[0]:>8} {a[0]:>8} {b[0]:>8} {c[0]:>9}")
    print("  " + "="*54)
    print("  " + ("전체 PASS" if ok_all else "일부 FAIL"))


# ── 진입점 ───────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", required=True, help="처리 날짜 (YYYY-MM-DD)")
    parser.add_argument("--from-master", action="store_true",
                        help="기준정보_마스터.xlsx A-L 데이터를 직접 사용 (raw 변환 생략)")
    args   = parser.parse_args()
    target = datetime.strptime(args.date, "%Y-%m-%d").date()

    results = process(target, from_master=args.from_master)

    _exp_d = _load_expected_from_final(args.date) or EXPECTED_BY_DATE.get(args.date, {})
    print(f"\n  [최종 결과]")
    for z, v in results.items():
        _e = _exp_d.get(z, 0)
        exp = _e["std"] if isinstance(_e, dict) else _e
        pct = abs(v['std_time_hr'] - exp) / exp * 100 if exp else 0
        print(f"    {z:<6}: 표준={v['std_time_hr']:.4f}hr  실적={v['act_time_hr']:.4f}hr"
              f"  (기준:{exp:.4f}  오차:{pct:.2f}%)")

    # 결과 JSON 저장 (비교용)
    import json
    out_path = BASE_DIR / f"data/temp/result_{args.date}.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump({z: {"std": round(v["std_time_hr"], 6), "act": round(v["act_time_hr"], 6)}
                   for z, v in results.items()}, f, ensure_ascii=False, indent=2)
    print(f"  [결과 저장] {out_path}")

    upsert(results, args.date)
    validate(results, args.date)


if __name__ == "__main__":
    main()
