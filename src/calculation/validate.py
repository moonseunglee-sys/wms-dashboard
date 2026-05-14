"""
툴파일 검증 시트(F_1 / I_1)와 자동계산 결과 비교
컬럼 위치는 F_1 기준 (I_1도 동일 템플릿):
  J(9)=작업자, I(8)=WAVE명, F(5)=LOCATION
  AB(27)=작업시간_min, AC(28)=작업시간_prime
  AS(44)=예상작업시간_min, AV(47)=작업소요시간_min
"""
import pandas as pd
import openpyxl


_COL = dict(worker=9, wave=8, loc=5, std=44, work=27, prime=28, adj=47)


def validate_sheet(
    detail_df: pd.DataFrame,
    tool_path: str,
    sheet_name: str = "I_1",
    target_worker: str | None = None,
    target_wave: str | None = None,
    threshold: float = 0.005,
    max_detail_rows: int = 60,
) -> None:
    """
    Args:
        detail_df:       calc_standard_times → calc_cumulative 결과
        tool_path:       툴 Excel 파일 경로
        sheet_name:      비교 시트명 (기본: "I_1")
        target_worker:   상세 행별 출력할 작업자 (None이면 첫 번째)
        target_wave:     상세 출력할 WAVE명 (None이면 첫 번째 WAVE)
        threshold:       차이 허용 기준(분), 초과 시 ✗ 표시
        max_detail_rows: 상세 출력 최대 행 수
    """
    ref_df = _load_ref_sheet(tool_path, sheet_name)
    if ref_df is None:
        return

    # ── 전체 작업자별 요약
    _print_summary(detail_df, ref_df, sheet_name, threshold)

    # ── 특정 작업자+WAVE 상세 비교
    workers = detail_df["작업자"].unique().tolist()
    if not workers:
        return

    w = target_worker if target_worker in workers else workers[0]
    our_w = detail_df[detail_df["작업자"] == w].reset_index(drop=True)
    ref_w = ref_df[ref_df["작업자"] == w].reset_index(drop=True)

    if ref_w.empty:
        print(f"\n[상세 스킵] I_1 시트에 '{w}' 데이터 없음")
        return

    # I_1에도 존재하는 WAVE 탐색
    ref_waves = set(ref_w["WAVE명"].unique())
    our_waves = our_w["WAVE명"].unique().tolist()

    if target_wave and target_wave in ref_waves:
        wv = target_wave
    else:
        matching = [wv for wv in our_waves if wv in ref_waves]
        if not matching:
            print(f"\n[상세 스킵] '{w}'의 WAVE가 I_1 시트와 일치하는 항목 없음")
            print(f"  자동화 WAVEs: {our_waves[:3]} ...")
            print(f"  I_1    WAVEs: {list(ref_waves)[:3]} ...")
            return
        wv = matching[0]

    our_wv = our_w[our_w["WAVE명"] == wv].reset_index(drop=True)
    ref_wv = ref_w[ref_w["WAVE명"] == wv].reset_index(drop=True)

    _print_detail(our_wv, ref_wv, w, wv, threshold, max_detail_rows)


# ─────────────────────────────────────────────
# 내부 함수
# ─────────────────────────────────────────────

def _load_ref_sheet(tool_path: str, sheet_name: str) -> pd.DataFrame | None:
    wb = openpyxl.load_workbook(tool_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        print(f"\n[검증 오류] '{sheet_name}' 시트 없음. 가용 시트: {wb.sheetnames}")
        wb.close()
        return None

    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        worker = row[_COL["worker"]]
        if worker is None:
            continue
        rows.append({
            "작업자":       str(worker),
            "WAVE명":       str(row[_COL["wave"]])  if row[_COL["wave"]]  is not None else "",
            "LOCATION":     row[_COL["loc"]],
            "ref_std":      float(row[_COL["std"]]   or 0),
            "ref_work":     float(row[_COL["work"]]  or 0),
            "ref_prime":    float(row[_COL["prime"]] or 0),
            "ref_adj":      float(row[_COL["adj"]]   or 0),
        })
    wb.close()

    if not rows:
        print(f"\n[검증 오류] '{sheet_name}' 시트가 비어 있습니다.")
        return None

    df = pd.DataFrame(rows)
    print(f"\n[{sheet_name}] 기준 시트 로드: {len(df)}행  작업자 {df['작업자'].nunique()}명")
    return df


def _print_summary(our: pd.DataFrame, ref: pd.DataFrame, sheet_name: str, threshold: float) -> None:
    workers = sorted(set(our["작업자"].unique()) | set(ref["작업자"].unique()))

    print(f"\n{'='*72}")
    print(f"[{sheet_name} 검증 요약]")
    print(f"{'작업자':<10} {'자동화행':>7} {'기준행':>6} {'비교행':>6} {'표준시간MAE':>11} {'작업소요MAE':>11} {'일치율':>7}")
    print(f"{'-'*72}")

    total_mae_std = total_mae_adj = total_n = 0

    for w in workers:
        o = our[our["작업자"] == w].reset_index(drop=True)
        r = ref[ref["작업자"] == w].reset_index(drop=True)

        if r.empty:
            print(f"  {w:<10} {len(o):>7}  {'없음':>6}")
            continue
        if o.empty:
            print(f"  {w:<10} {'없음':>7}  {len(r):>6}")
            continue

        n = min(len(o), len(r))
        mae_std = (o["예상작업시간_min"].iloc[:n] - r["ref_std"].iloc[:n]).abs().mean()
        mae_adj = (o["작업소요시간_min"].iloc[:n] - r["ref_adj"].iloc[:n]).abs().mean()
        match_r = ((o["예상작업시간_min"].iloc[:n] - r["ref_std"].iloc[:n]).abs() < threshold).mean() * 100

        total_mae_std += mae_std * n
        total_mae_adj += mae_adj * n
        total_n += n

        print(f"  {w:<10} {len(o):>7} {len(r):>6} {n:>6}  {mae_std:>10.4f}분  {mae_adj:>10.4f}분  {match_r:>6.1f}%")

    if total_n > 0:
        print(f"{'-'*72}")
        print(f"  {'전체':<10} {len(our):>7} {len(ref):>6} {total_n:>6}  {total_mae_std/total_n:>10.4f}분  {total_mae_adj/total_n:>10.4f}분")


def _print_detail(our: pd.DataFrame, ref: pd.DataFrame, worker: str, wave: str,
                  threshold: float, max_rows: int) -> None:
    n = min(len(our), len(ref), max_rows)
    print(f"\n{'='*80}")
    print(f"[상세 비교] 작업자: {worker}  WAVE: {wave}")
    print(f"  자동화 {len(our)}행  기준시트 {len(ref)}행  출력 {n}행")
    print(f"\n  {'LOCATION':<18} {'기준예상':>8} {'자동화':>8} {'차이':>7} | {'기준작업소요':>10} {'자동화':>8} {'차이':>7}")
    print(f"  {'-'*78}")

    total_ds = total_da = 0.0
    for j in range(n):
        ds = our.loc[j, "예상작업시간_min"] - ref.loc[j, "ref_std"]
        da = our.loc[j, "작업소요시간_min"]  - ref.loc[j, "ref_adj"]
        total_ds += abs(ds)
        total_da += abs(da)
        flag = "O" if abs(ds) < threshold else "X"
        loc  = str(ref.loc[j, "LOCATION"] or "")
        print(f"  {flag} {loc:<17} "
              f"{ref.loc[j,'ref_std']:>8.4f} {our.loc[j,'예상작업시간_min']:>8.4f} {ds:>+7.4f} | "
              f"{ref.loc[j,'ref_adj']:>10.4f} {our.loc[j,'작업소요시간_min']:>8.4f} {da:>+7.4f}")

    ref_sum = ref["ref_std"].iloc[:n].sum()
    our_sum = our["예상작업시간_min"].iloc[:n].sum()
    print(f"  {'-'*78}")
    print(f"  {'합계':<18} {ref_sum:>8.4f} {our_sum:>8.4f} {our_sum-ref_sum:>+7.4f}")
    print(f"  표준시간 MAE: {total_ds:.4f}분  |  작업소요시간 MAE: {total_da:.4f}분")
