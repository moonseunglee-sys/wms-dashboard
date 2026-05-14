"""
기준값 로드 — 양지센터_피킹_가동율.xlsx 에서 이동시간/피킹시간/WAVE간시간 파싱
"""
from pathlib import Path
import openpyxl


def load_reference_tables(tool_path: str | Path) -> dict:
    wb = openpyxl.load_workbook(tool_path, read_only=True, data_only=True)
    ref = {}

    ref["zone_dist"]       = _load_zone_dist(wb)
    ref["rack_dist_efhikl"], ref["rack_dist_abcdj"] = _load_rack_dist(wb)
    ref["loc_dist"]        = _load_loc_dist(wb)
    ref["tier_time"]       = _load_tier_time(wb)
    _load_pick_time(wb, ref)      # pick_time, barcode_item, barcode_loc, start_loc, end_loc
    _load_wave_time(wb, ref)      # pallet_prep, labeling

    wb.close()
    return ref


# ── ABCD: zone간 이동시간 (초)
def _load_zone_dist(wb) -> dict:
    ws = wb["ABCD"]
    rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
    zone_labels = [v for v in rows[1] if v is not None]
    dist = {}
    for row in rows[2:]:
        if row[0] is None:
            continue
        fz = row[0]
        for j, tz in enumerate(zone_labels, 1):
            if j < len(row) and isinstance(row[j], (int, float)):
                dist[(fz, tz)] = float(row[j])
    return dist


# ── 101_EFHIKL / 101_ABCDJ: rack간 이동시간 (초)
def _load_rack_dist(wb) -> tuple[dict, dict]:
    result = {}
    for sheet_name in ("101_EFHIKL", "101_ABCDJ"):
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[1]
        rack_labels = [int(v) for v in header[2:] if isinstance(v, (int, float))]
        store = {}
        for row in rows[2:]:
            if not isinstance(row[1], (int, float)):
                continue
            fr = int(row[1])
            for j, tr in enumerate(rack_labels):
                col_idx = j + 2
                if col_idx < len(row) and isinstance(row[col_idx], (int, float)):
                    store[(fr, tr)] = float(row[col_idx])
        result[sheet_name] = store
    return result["101_EFHIKL"], result["101_ABCDJ"]


# ── 1_24: location간 이동시간 (초)
def _load_loc_dist(wb) -> dict:
    ws = wb["1_24"]
    rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
    loc_labels = [int(v) for v in rows[1] if isinstance(v, (int, float))]
    dist = {}
    for row in rows[2:]:
        if not isinstance(row[0], (int, float)):
            continue
        fl = int(row[0])
        for j, tl in enumerate(loc_labels, 1):
            if j < len(row) and isinstance(row[j], (int, float)):
                dist[(fl, tl)] = float(row[j])
    return dist


# ── 1_10: 단 높이별 피킹 소요시간 (분)
def _load_tier_time(wb) -> dict:
    ws = wb["1_10"]
    tier_time = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is not None and isinstance(row[1], (int, float)):
            tier_time[row[0]] = float(row[1])
    return tier_time


# ── 피킹시간 시트: zone별 피킹시간, 바코드 스캔시간, 시작/끝 로케이션 테이블
def _load_pick_time(wb, ref: dict) -> None:
    ws = wb["피킹시간"]
    rows = list(ws.iter_rows(values_only=True))

    # 행3~: A열=zone, B열=피킹시간(분)
    pick_time = {}
    for row in rows[2:]:
        if row[0] is not None and isinstance(row[1], (int, float)):
            pick_time[row[0]] = float(row[1])
    ref["pick_time"] = pick_time

    # I열(8) = 품목 바코드 스캔(분), M열(12) = loc 바코드 스캔(분)
    ref["barcode_item"] = float(rows[2][8])  if isinstance(rows[2][8],  (int, float)) else 0.13333
    ref["barcode_loc"]  = float(rows[2][12]) if isinstance(rows[2][12], (int, float)) else 0.13333

    # 시작 위치 테이블: P열(15) 키 → (Q zone, R rack, S loc)
    # 끝 위치 테이블:   U열(20) 키 → (V zone, W rack, X loc)
    start_loc_table = {}
    end_loc_table   = {}
    for row in rows[2:]:
        if row[0] is None:
            continue
        p_key = row[15]
        if p_key is not None:
            q, r_val, s = row[16], row[17], row[18]
            if q is not None:
                start_loc_table[p_key] = (
                    q,
                    int(r_val) if isinstance(r_val, (int, float)) else None,
                    int(s)     if isinstance(s,     (int, float)) else None,
                )
        u_key = row[20]
        if u_key is not None:
            v, w, x = row[21], row[22], row[23]
            if v is not None:
                end_loc_table[u_key] = (
                    v,
                    int(w) if isinstance(w, (int, float)) else None,
                    int(x) if isinstance(x, (int, float)) else None,
                )
    ref["start_loc"] = start_loc_table
    ref["end_loc"]   = end_loc_table


# ── WAVE간 시간: 공파렛트 준비 + 출고지역별 라벨링/복귀 (분)
def _load_wave_time(wb, ref: dict) -> None:
    ws = wb["WAVE간 시간"]
    rows = list(ws.iter_rows(values_only=True))

    pallet_prep = {}
    for row in rows[2:5]:
        if row[0] is not None and isinstance(row[1], (int, float)):
            pallet_prep[row[0]] = float(row[1])
    ref["pallet_prep"] = pallet_prep

    # 출고지역별 라벨링+복귀: key_col=zone, val_col=시간(분)
    region_col = {
        "지방": (9,  10),
        "소액": (15, 16),
        "경인": (22, 23),
        "수출": (29, 30),
    }
    labeling = {r: {} for r in region_col}
    for region, (key_col, val_col) in region_col.items():
        for row in rows[3:]:
            if len(row) > key_col and isinstance(row[key_col], str):
                zone_key = row[key_col]
                val = row[val_col] if len(row) > val_col and isinstance(row[val_col], (int, float)) else 0.0
                labeling[region][zone_key] = float(val)
    ref["labeling"] = labeling
